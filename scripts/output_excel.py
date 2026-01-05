#!/usr/bin/env python3

# ---------- START IMPORT PACKAGES ----------
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
# ---------- END IMPORT PACKAGES ----------

# ---------- START DECLARE ROOT FOLDER ----------
ROOT = Path(__file__).resolve().parents[1]
# ---------- END DECLARE ROOT FOLDER ----------

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "1. SBOM"

headers = [
    "#",
    "System / Application",
    "Purpose / Usage",
    "URL",
    "Services Mode",
    "Target Customer",
    "Software Component",
    "Third-party Modules",
    "External APIs or Services",
    "Critical Level",
    "Data Category",
    "Is the application/system currently in use?",
    "Application/System Developer",
    "Vendor's Name",
    "Does the agency have expertise?",
    "Does the agency have a special budget allocation?",
    "Link to CBOM",
]

n_cols = len(headers)
last_col = get_column_letter(n_cols)  # Q

# A1 title
ws.merge_cells(f"A1:{last_col}1")
ws["A1"] = "Table 1 – Software Bill of Materials (SBOM)"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# A2 description
ws.merge_cells(f"A2:{last_col}2")
ws["A2"] = ("This table lists software components used in systems identified in Table 0, focusing on cryptographic usage within libraries and software packages.")
ws["A2"].font = Font(size=11, italic=True)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Header row at A4
header_row = 4
header_fill = PatternFill("solid", fgColor="F4B183")  # orange header like your screenshot
header_font = Font(bold=True, size=10)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col_idx, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = header_alignment
    c.border = border

# Row heights
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 36
ws.row_dimensions[4].height = 48

# Freeze panes below headers
ws.freeze_panes = "A5"

# Column widths (A4 landscape-friendly)
widths = {
    1: 4,  2: 20, 3: 22, 4: 22, 5: 14, 6: 16, 7: 18, 8: 18, 9: 20,
    10: 12, 11: 14, 12: 22, 13: 20, 14: 16, 15: 18, 16: 22, 17: 16
}
for idx in range(1, n_cols + 1):
    ws.column_dimensions[get_column_letter(idx)].width = widths.get(idx, 18)

# Page setup: A4, fit to 1 page width
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.scale = None
ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)

# Print titles and area
ws.print_title_rows = "1:4"
ws.print_area = f"A1:{last_col}200"

wb.save((ROOT / "output" / "sbom_template.xlsx").resolve())
