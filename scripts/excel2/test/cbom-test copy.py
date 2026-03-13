#!/usr/bin/env python3
"""
CBOM + SBOM (Syft) + Vulnerability (Grype) XLSX generator

Inputs:
  - cbom.json                 (Semgrep JSON with metadata.cbom fields)
  - /mnt/data/sbom.json       (Syft JSON)
  - /mnt/data/grype.json      (Grype JSON)

Output:
  - cbom_output.xlsx

Sheets:
  1) CBOM
  2) CBOM_Summary
  3) SBOM_Grype   (example columns you provided)

Header styling:
  - Light gray header fill + bold + centered + thin borders (all sheets)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd


# ====== CONFIG ======
INPUT_CBOM_JSON = Path(r"cbom.json")
INPUT_SBOM_JSON = Path(r"sbom.json")
INPUT_GRYPE_JSON = Path(r"grype.json")
OUT_XLSX = Path("cbom_output.xlsx")
# ====================


CBOM_COLS = [
    "# (CBOM)",
    "System / Application",
    "Cryptographic Function",
    "Algorithm Used",
    "Library / Module",
    "File / Location",
    "Key Length",
    "Purpose / Usage",
    "Crypto-Agility Support",
]

SBOM_SHEET_COLS = [
    "#",
    "System / Application",
    "Purpose / Usage",
    "URL",
    "Services Mode",
    "Target Customer",
    "Software Component",
    "Vulnerability Status",
    "Third-party Modules",
    "External APIs or Services",
    "Critical Level",
    "Data Category",
    "Is the application/system currently in use?",
    "Application/System Developer",
    "Vendor's Name",
    "Does the agency have expertise?",
    "Does the agency have a special budget allocation?",
    "Link to CBOM",
]


# -------------------- GENERAL HELPERS --------------------
def coalesce(*vals: Any, default: str = "Unknown") -> str:
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return str(v)
    return default


def _bullet_list(items: List[str], top_n: int = 10, header: Optional[str] = None) -> str:
    items = [x for x in items if x and str(x).strip()]
    if not items:
        return "Unknown"

    uniq: List[str] = []
    seen = set()
    for x in items:
        key = x.strip()
        if key in seen:
            continue
        seen.add(key)
        uniq.append(key)

    shown = uniq[:top_n]
    remaining = max(0, len(uniq) - len(shown))

    lines: List[str] = []
    if header:
        lines.append(header)
    lines.extend([f"• {x}" for x in shown])
    if remaining > 0:
        lines.append(f"... and {remaining} more.")
    return "\n".join(lines)


def _best_system_name_from_path(p: str) -> str:
    try:
        parts = Path(p).parts
        if not parts:
            return "Unknown"
        # last folder name
        return parts[-1] or "Unknown"
    except Exception:
        return "Unknown"


def _severity_rank(sev: str) -> int:
    s = (sev or "").upper()
    return {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}.get(s, 0)


def _max_severity(sevs: List[str]) -> str:
    best = ""
    best_rank = 0
    for s in sevs:
        r = _severity_rank(s)
        if r > best_rank:
            best_rank = r
            best = s
    return best.upper() if best_rank > 0 else "None"


# -------------------- SEMGREP CBOM BUILDERS --------------------
def guess_system_application(file_path: str) -> str:
    p = Path(file_path)
    parts = list(p.parts)

    if "testing" in parts:
        idx = parts.index("testing")
        if idx + 1 < len(parts):
            return parts[idx + 1]

    if "GitHub" in parts:
        idx = parts.index("GitHub")
        if idx + 1 < len(parts):
            return parts[idx + 1]

    if p.parent.name:
        return p.parent.name

    return "Unknown"


def extract_line_number(result: Dict[str, Any]) -> Optional[int]:
    start = result.get("start") or {}
    line = start.get("line")
    if isinstance(line, int):
        return line

    extra = result.get("extra") or {}
    line2 = extra.get("line")
    if isinstance(line2, int):
        return line2

    return None


def format_file_location_with_line(path: str, line: Optional[int]) -> str:
    if not path:
        return ""
    if line is None:
        return path
    return f"{path} ({line})"


def add_cbom_numbering(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    df = df.copy()
    groups = {name: i + 1 for i, name in enumerate(df[group_col].dropna().unique())}
    group_id = df[group_col].map(groups).fillna(0).astype(int)
    within = df.groupby(group_col, dropna=False).cumcount() + 1
    df.insert(0, "# (CBOM)", group_id.astype(str) + "." + within.astype(str))
    return df


def build_cbom_rows(semgrep: Dict[str, Any]) -> List[Dict[str, Any]]:
    results = semgrep.get("results", [])
    rows: List[Dict[str, Any]] = []

    for r in results:
        path = r.get("path", "") or ""
        line = extract_line_number(r)
        file_location = format_file_location_with_line(path, line)

        extra = r.get("extra", {}) or {}
        md = extra.get("metadata", {}) or {}
        cbom = md.get("cbom", {}) or {}

        rows.append(
            {
                "System / Application": coalesce(md.get("system_application"), default=guess_system_application(path)),
                "Cryptographic Function": coalesce(cbom.get("cryptographic_function"), md.get("cryptographic_function")),
                "Algorithm Used": coalesce(cbom.get("algorithm_used"), md.get("algorithm_name")),
                "Library / Module": coalesce(cbom.get("library_module"), md.get("library")),
                "File / Location": file_location,
                "Key Length": coalesce(cbom.get("key_length")),
                "Purpose / Usage": coalesce(cbom.get("purpose_usage"), md.get("purpose")),
                "Crypto-Agility Support": coalesce(md.get("crypto_agility"), cbom.get("crypto_agility")),
            }
        )
    return rows


def summarize_top_locations(file_locations: List[str], top_n: int = 3) -> str:
    seen = set()
    uniq: List[str] = []
    for loc in file_locations:
        loc = (loc or "").strip()
        if not loc or loc in seen:
            continue
        seen.add(loc)
        uniq.append(loc)

    shown = uniq[:top_n]
    remaining = max(0, len(uniq) - len(shown))

    if not shown:
        return "Unknown"

    bullets = "\n".join([f"• {x}" for x in shown])
    if remaining > 0:
        bullets += f"\n... and {remaining} more files."
    return bullets


def build_cbom_summary(df_full_no_number: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for crypto_func, g in df_full_no_number.groupby("Cryptographic Function", dropna=False, sort=True):
        rows.append(
            {
                "System / Application": g["System / Application"].iloc[0],
                "Cryptographic Function": crypto_func if str(crypto_func).strip() else "Unknown",
                "Algorithm Used": g["Algorithm Used"].iloc[0],
                "Library / Module": g["Library / Module"].iloc[0],
                "File / Location": summarize_top_locations(g["File / Location"].tolist(), top_n=3),
                "Key Length": g["Key Length"].iloc[0],
                "Purpose / Usage": g["Purpose / Usage"].iloc[0],
                "Crypto-Agility Support": g["Crypto-Agility Support"].iloc[0],
            }
        )

    df_sum = pd.DataFrame(rows)
    if not df_sum.empty:
        df_sum.sort_values(["Cryptographic Function", "System / Application"], inplace=True, kind="stable")
        df_sum = add_cbom_numbering(df_sum, group_col="Cryptographic Function")
        df_sum = df_sum[CBOM_COLS]
    else:
        df_sum = pd.DataFrame(columns=CBOM_COLS)
    return df_sum


# -------------------- SYFT SBOM PARSER --------------------
def parse_syft_sbom(sbom: Dict[str, Any]) -> Dict[str, Any]:
    """
    Returns:
      {
        "system_name": str,
        "third_party_modules": List[str],   # "name (version)" strings
        "developer_guess": str,
      }
    """
    src = sbom.get("source") or {}
    src_name = coalesce(src.get("name"), default="")
    system_name = _best_system_name_from_path(src_name) if src_name else coalesce(sbom.get("name"), default="Unknown")

    artifacts = sbom.get("artifacts") or []
    mods: List[str] = []
    for a in artifacts:
        name = a.get("name")
        ver = a.get("version")
        typ = a.get("type")
        if name and ver:
            mods.append(f"{name} ({ver})")
        elif name:
            mods.append(f"{name} ({typ})" if typ else name)

    # crude "developer" guess: try to detect wordpress or common vendors
    developer_guess = "Unknown Vendor"
    if any("wordpress" in (m.lower()) for m in mods):
        developer_guess = "WordPress Foundation"

    return {
        "system_name": system_name,
        "third_party_modules": mods,
        "developer_guess": developer_guess,
        "source_name": src_name,
    }


# -------------------- GRYPE PARSER --------------------
def parse_grype(grype: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract a summary list of vulnerabilities:
      [
        "package: [SEVERITY] CVE-XXXX-YYYY"
      ]
    and return highest severity.
    """
    matches = grype.get("matches") or []
    vul_lines: List[str] = []
    sevs: List[str] = []

    for m in matches:
        vuln = (m.get("vulnerability") or {})
        artifact = (m.get("artifact") or {})

        cve = coalesce(vuln.get("id"), default="UNKNOWN")
        sev = coalesce(vuln.get("severity"), default="UNKNOWN").upper()

        pkg = coalesce(artifact.get("name"), default="package")
        # optional: add version if present
        ver = artifact.get("version")
        pkg_disp = f"{pkg} {ver}".strip() if ver else pkg

        vul_lines.append(f"{pkg_disp}: [{sev}] {cve}")
        sevs.append(sev)

    critical_level = _max_severity(sevs)

    return {
        "vuln_lines": vul_lines,
        "critical_level": critical_level,
        "match_count": len(matches),
    }


# -------------------- BUILD SBOM SHEET --------------------
def build_sbom_sheet(
    df_cbom_full: pd.DataFrame,
    sbom: Dict[str, Any],
    grype: Dict[str, Any],
) -> pd.DataFrame:
    sy = parse_syft_sbom(sbom)
    gy = parse_grype(grype)

    system_name = sy["system_name"]

    # Try to link CBOM entries for this system name (best-effort)
    cbom_systems = set(df_cbom_full["System / Application"].dropna().astype(str).tolist())
    if system_name in cbom_systems:
        link_cbom = f"See CBOM entries for {system_name}"
    else:
        # fallback: just point to CBOM sheet
        link_cbom = "See CBOM sheet"

    software_component = "\n".join(
        [
            "[CODE LEVEL]",
            f"App/System: {system_name}",
            f"SBOM Source: {sy.get('source_name') or 'Unknown'}",
        ]
    )

    vuln_status = "No known vulnerabilities found" if gy["match_count"] == 0 else _bullet_list(
        gy["vuln_lines"],
        top_n=40,
        header="[ISSUES FOUND]",
    )

    third_party = _bullet_list(
        sy["third_party_modules"],
        top_n=60,
        header="[DEPENDENCIES]",
    )

    # Build row (leave unknowns blank unless you want defaults)
    row = {
        "#": 1,
        "System / Application": system_name,
        "Purpose / Usage": "",
        "URL": "",
        "Services Mode": "",
        "Target Customer": "",
        "Software Component": software_component,
        "Vulnerability Status": vuln_status,
        "Third-party Modules": third_party,
        "External APIs or Services": "",
        "Critical Level": gy["critical_level"],
        "Data Category": "",
        "Is the application/system currently in use?": "",
        "Application/System Developer": sy["developer_guess"],
        "Vendor's Name": "",
        "Does the agency have expertise?": "",
        "Does the agency have a special budget allocation?": "",
        "Link to CBOM": link_cbom,
    }

    df = pd.DataFrame([row], columns=SBOM_SHEET_COLS)
    return df


# -------------------- EXCEL STYLING --------------------
def style_header(ws) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # light gray
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border


def autosize_columns(ws) -> None:
    from openpyxl.utils import get_column_letter

    max_width_cap = 90
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, max((len(x) for x in v.splitlines()), default=0))
        width = min(max_width_cap, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# -------------------- MAIN --------------------
def main() -> None:
    # --- CBOM ---
    with INPUT_CBOM_JSON.open("r", encoding="utf-8") as f:
        semgrep = json.load(f)

    df_full = pd.DataFrame(build_cbom_rows(semgrep))
    if not df_full.empty:
        df_full.sort_values(
            ["Cryptographic Function", "System / Application", "File / Location", "Algorithm Used"],
            inplace=True,
            kind="stable",
            na_position="last",
        )

    df_full = add_cbom_numbering(df_full, group_col="System / Application")
    df_full = df_full[CBOM_COLS]
    df_sum = build_cbom_summary(df_full.drop(columns=["# (CBOM)"], errors="ignore"))

    # --- SBOM + Grype ---
    with INPUT_SBOM_JSON.open("r", encoding="utf-8") as f:
        sbom = json.load(f)
    with INPUT_GRYPE_JSON.open("r", encoding="utf-8") as f:
        grype = json.load(f)

    df_sbom = build_sbom_sheet(df_full, sbom, grype)

    # --- Write XLSX ---
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df_full.to_excel(writer, index=False, sheet_name="CBOM")
        df_sum.to_excel(writer, index=False, sheet_name="CBOM_Summary")
        df_sbom.to_excel(writer, index=False, sheet_name="SBOM_Grype")

        for name in ["CBOM", "CBOM_Summary", "SBOM_Grype"]:
            ws = writer.sheets[name]
            ws.freeze_panes = "A2"
            style_header(ws)
            autosize_columns(ws)

    print(f"[OK] Wrote: {OUT_XLSX.resolve()}")
    print(f"[OK] Rows: CBOM={len(df_full)} | CBOM_Summary={len(df_sum)} | SBOM_Grype={len(df_sbom)}")


if __name__ == "__main__":
    main()