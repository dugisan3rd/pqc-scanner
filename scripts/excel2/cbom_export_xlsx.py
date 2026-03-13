import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INPUT_FILE = "semgrep_cbom.json"   # <-- rename to your json filename
OUTPUT_FILE = "cbom_output.xlsx"
SHEET_NAME = "2_CBOM"

HEADERS = [
    "# (CBOM)",
    "System / Application",
    "Cryptographic Function",
    "Algorithm Used",
    "Algorithm Category",
    "Library / Module",
    "File / Location",
    "Key Length",
    "Purpose / Usage",
    "Crypto-Agility Support",
]

IGNORE_FOLDERS = {
    "src", "app", "lib", "dist", "build", "vendor", "node_modules",
    "venv", ".venv", "test", "tests", "__tests__", "fixtures", "docs",
    "target", "bin", "obj", "out"
}

def detect_system_name(file_path: str) -> str:
    if not file_path:
        return "Unknown"
    parts = file_path.replace("\\", "/").split("/")
    for p in parts:
        if p and p.lower() not in IGNORE_FOLDERS:
            return p
    return "Unknown"

def auto_crypto_agility(cbom: dict) -> str:
    algo = (cbom.get("algorithm_used") or "").lower()
    if "md5" in algo or "sha1" in algo:
        return "Low"
    if "rsa" in algo or "ecd" in algo or "ecdh" in algo or "ecdsa" in algo:
        return "Medium"
    if any(pqc in algo for pqc in ["kyber", "dilithium", "falcon", "sphincs"]):
        return "High"
    return ""

def normalize_code_snippet(s: str, max_len: int = 200) -> str:
    """
    Convert multi-line match to single-line, remove excessive whitespace,
    and cap length so XLSX stays readable.
    """
    if not s:
        return ""
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = " | ".join(line.strip() for line in s.split("\n") if line.strip())
    if len(s) > max_len:
        s = s[:max_len].rstrip() + "…"
    return s

def build_file_location(finding: dict) -> str:
    """
    File / Location format:
    <path> (Lstart[-Lend]) [<matched code>]
    """
    path = finding.get("path", "") or ""
    start_line = (finding.get("start") or {}).get("line")
    end_line = (finding.get("end") or {}).get("line")
    code = normalize_code_snippet((finding.get("extra") or {}).get("lines", ""))

    if start_line and end_line and start_line != end_line:
        line_part = f"(L{start_line}-L{end_line})"
    elif start_line:
        line_part = f"(L{start_line})"
    else:
        line_part = ""

    if code:
        return f"{path} {line_part} [{code}]".strip()
    return f"{path} {line_part}".strip()

def auto_adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 90)

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"[!] Input file not found: {INPUT_FILE}")
        return

    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    findings = data.get("results", [])

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, finding in enumerate(findings, start=1):
        metadata = (finding.get("extra") or {}).get("metadata", {}) or {}
        cbom = metadata.get("cbom", {}) or {}

        file_location = build_file_location(finding)
        system_name = cbom.get("system_application") or detect_system_name(finding.get("path", ""))

        crypto_agility = metadata.get("crypto_agility_support") or cbom.get("crypto_agility_support") or auto_crypto_agility(cbom)

        ws.append([
            i,
            system_name,
            cbom.get("cryptographic_function", ""),
            cbom.get("algorithm_used", ""),
            cbom.get("algorithm_category", ""),
            cbom.get("library_module", ""),
            file_location,
            cbom.get("key_length", ""),
            cbom.get("purpose_usage", ""),
            crypto_agility,
        ])

    auto_adjust_columns(ws)
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)

    print(f"[+] Exported: {OUTPUT_FILE}")
    print(f"[+] Sheet: {SHEET_NAME}")
    print(f"[+] Rows: {len(findings)}")

if __name__ == "__main__":
    main()