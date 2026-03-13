import json
import os
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
INPUT_FILE = "semgrep_cbom.json"   # <-- rename your semgrep JSON filename (or change this)
OUTPUT_FILE = "cbom_output.xlsx"

SHEET_CBOM = "2_CBOM"
SHEET_RISK = "3_RiskRegister"

CBOM_HEADERS = [
    "# (CBOM)",
    "System / Application",
    "Cryptographic Function",
    "Algorithm Used",
    "Algorithm Category",
    "Library / Module",
    "File / Location",
    "Key Length",
    "Purpose / Usage",
    "Crypto-Agility Support",
]

RISK_HEADERS = [
    "#",
    "Nama Sistem/ Perkakasan/Perisian",
    "Jenis Aset",
    "Algoritma Kriptografi",
    "Kegunaan Algoritma Kriptografi",
    "Tahap Kritikal",
    "Risiko",
    "Pemilik Risiko",
]

IGNORE_FOLDERS = {
    "src", "app", "lib", "dist", "build", "vendor", "node_modules",
    "venv", ".venv", "test", "tests", "__tests__", "fixtures", "docs",
    "target", "bin", "obj", "out", ".git"
}

RISK_OWNER_DEFAULT = "IT Security Team"

# =========================
# HELPERS
# =========================
def detect_system_name(file_path: str) -> str:
    # Best-effort system name inference from file path (no hardcoding).
    if not file_path:
        return "Unknown"
    parts = file_path.replace("\\", "/").split("/")
    for p in parts:
        if p and p.lower() not in IGNORE_FOLDERS:
            return p
    return "Unknown"

def normalize_code_snippet(s: str, max_len: int = 220) -> str:
    if not s:
        return ""
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = " | ".join(line.strip() for line in s.split("\n") if line.strip())
    if len(s) > max_len:
        s = s[:max_len].rstrip() + "…"
    return s

def build_file_location(finding: dict) -> str:
    # <path> (Lstart[-Lend]) [<matched code>]
    path = finding.get("path", "") or ""
    start_line = (finding.get("start") or {}).get("line")
    end_line = (finding.get("end") or {}).get("line")
    code = normalize_code_snippet((finding.get("extra") or {}).get("lines", ""))

    if start_line and end_line and start_line != end_line:
        line_part = f"(L{start_line}-L{end_line})"
    elif start_line:
        line_part = f"(L{start_line})"
    else:
        line_part = ""

    if code:
        return f"{path} {line_part} [{code}]".strip()
    return f"{path} {line_part}".strip()

def auto_crypto_agility(cbom: dict) -> str:
    algo = (cbom.get("algorithm_used") or "").lower()
    if "md5" in algo or "sha1" in algo:
        return "Low"
    if any(x in algo for x in ["rsa", "ecd", "ecdh", "ecdsa", "dsa", "dh"]):
        return "Medium"
    if any(pqc in algo for pqc in ["kyber", "dilithium", "falcon", "sphincs", "frodokem", "hqc"]):
        return "High"
    return ""

def autosize_columns(ws, max_width=90):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)

def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def classify_asset_type(cbom: dict, finding: dict) -> str:
    lib = (cbom.get("library_module") or "").lower()
    cat = (cbom.get("algorithm_category") or "").lower()
    path = (finding.get("path") or "").lower()

    if any(x in path for x in ["/etc/", ".yml", ".yaml", ".conf", ".ini", ".properties", ".toml", ".env"]):
        return "Configuration"
    if any(x in lib for x in ["openssl", "crypto", "webcrypto", "bouncycastle", "cryptography", "tls"]):
        return "Software Component"
    if "application" in cat or "source" in cat or any(x in path for x in [".php", ".js", ".ts", ".py", ".java", ".go", ".cs", ".rb"]):
        return "Application Code"
    return "Software"

def classify_criticality_and_risk(cbom: dict) -> tuple[str, str]:
    algo = (cbom.get("algorithm_used") or "").lower()
    cat = (cbom.get("algorithm_category") or "").lower()
    func = (cbom.get("cryptographic_function") or "").lower()

    # TLS / verification disabled or legacy TLS (not purely PQC, still critical)
    if "tls" in cat or "secure channel" in func:
        return ("Kritikal", "TLS Misconfiguration: Potential interception / downgrade risk")

    # Asymmetric families (PQC at-risk)
    if any(x in algo for x in ["rsa", "ecd", "ecdh", "ecdsa", "dsa", "dh", "ed25519", "ed448"]):
        return ("Kritikal", "PQC Vulnerability: Encryption/Signature broken by Quantum Computer")

    # Explicit weak hashes
    if "md5" in algo or "sha1" in algo:
        return ("Tinggi", "PQC Vulnerability: Brute-force resistance reduced by 50% (also weak legacy hash)")

    # Symmetric / hashing general
    if any(x in cat for x in ["hash", "symmetric", "encryption", "mac"]) or "hash" in func or "encryption" in func:
        return ("Tinggi", "PQC Vulnerability: Brute-force resistance reduced by 50%")

    return ("Sederhana", "PQC Readiness: Cryptographic usage should be reviewed for crypto-agility")

def normalize_usage(cbom: dict) -> str:
    return cbom.get("purpose_usage") or cbom.get("cryptographic_function") or ""

def row_key_for_dedupe(system_name: str, algo: str, usage: str, asset_type: str) -> str:
    return f"{system_name}||{asset_type}||{algo}||{usage}".lower().strip()

# =========================
# MAIN
# =========================
def main():
    if not os.path.exists(INPUT_FILE):
        print(f"[!] Input file not found: {INPUT_FILE}")
        print("[!] Tip: rename your semgrep output JSON to semgrep_cbom.json or update INPUT_FILE.")
        return

    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    findings = data.get("results", [])

    wb = Workbook()

    # -------------------------
    # Sheet 2_CBOM
    # -------------------------
    ws_cbom = wb.active
    ws_cbom.title = SHEET_CBOM
    ws_cbom.append(CBOM_HEADERS)
    style_header(ws_cbom)

    cbom_rows = []
    for i, finding in enumerate(findings, start=1):
        metadata = (finding.get("extra") or {}).get("metadata", {}) or {}
        cbom = (metadata.get("cbom") or {}) if isinstance(metadata.get("cbom"), dict) else {}

        file_path = finding.get("path", "")
        file_location = build_file_location(finding)

        system_name = cbom.get("system_application") or detect_system_name(file_path)
        crypto_agility = (
            metadata.get("crypto_agility_support")
            or cbom.get("crypto_agility_support")
            or auto_crypto_agility(cbom)
        )

        cbom_rows.append([
            i,
            system_name,
            cbom.get("cryptographic_function", ""),
            cbom.get("algorithm_used", ""),
            cbom.get("algorithm_category", ""),
            cbom.get("library_module", ""),
            file_location,
            cbom.get("key_length", ""),
            cbom.get("purpose_usage", ""),
            crypto_agility,
        ])

    for row in cbom_rows:
        ws_cbom.append(row)

    ws_cbom.freeze_panes = "A2"
    autosize_columns(ws_cbom, max_width=110)

    # -------------------------
    # Sheet 3_RiskRegister (derived from CBOM)
    # -------------------------
    ws_risk = wb.create_sheet(SHEET_RISK)
    ws_risk.append(RISK_HEADERS)
    style_header(ws_risk)

    dedup = OrderedDict()  # keep order and dedupe repeated hits
    for finding in findings:
        metadata = (finding.get("extra") or {}).get("metadata", {}) or {}
        cbom = (metadata.get("cbom") or {}) if isinstance(metadata.get("cbom"), dict) else {}

        file_path = finding.get("path", "")
        system_name = cbom.get("system_application") or detect_system_name(file_path)
        algo = cbom.get("algorithm_used", "") or ""
        usage = normalize_usage(cbom)
        asset_type = classify_asset_type(cbom, finding)

        if not algo and not usage:
            continue

        key = row_key_for_dedupe(system_name, algo, usage, asset_type)
        if key not in dedup:
            criticality, risk_text = classify_criticality_and_risk(cbom)
            dedup[key] = {
                "system": system_name,
                "asset_type": asset_type,
                "algo": algo,
                "usage": usage,
                "criticality": criticality,
                "risk": risk_text,
                "owner": metadata.get("risk_owner") or RISK_OWNER_DEFAULT,
            }

    for idx, item in enumerate(dedup.values(), start=1):
        ws_risk.append([
            idx,
            item["system"],
            item["asset_type"],
            item["algo"],
            item["usage"],
            item["criticality"],
            item["risk"],
            item["owner"],
        ])

    ws_risk.freeze_panes = "A2"
    autosize_columns(ws_risk, max_width=90)

    wb.save(OUTPUT_FILE)
    print(f"[+] Exported: {OUTPUT_FILE}")
    print(f"[+] Sheets: {SHEET_CBOM}, {SHEET_RISK}")
    print(f"[+] 2_CBOM rows: {len(cbom_rows)}")
    print(f"[+] 3_RiskRegister rows (deduped): {len(dedup)}")

if __name__ == "__main__":
    main()
