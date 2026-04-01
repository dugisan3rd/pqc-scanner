#!/usr/bin/env python3
"""Generate PQC workbook matching BUKUKERJA_BENGKEL MIGRASI PQC 2025.xlsx format.

Uses the reference template from bin/pqc_migration/ as a base, then populates
the 4 data sheets (SBOM, CBOM, RiskRegister, RiskAssessment) with scan results.
Static sheets (0_Inventory, 5_RiskMatrix, 6_ProtocolCryptoMap, 00_ReadMe) are
preserved as-is from the template.

Inputs:
  1) Syft SBOM JSON
  2) Grype vulnerability JSON
  3) Semgrep CBOM JSON
  4) mini_pqc JSON (reserved)

Usage:
  python3 pqc_export_full.py --sbom 1_syft.json --grype 2_grype.json \
      --cbom 4_semgrep_cbom.json --mini 3_mini_pqc.json --out pqc_report.xlsx
"""

import json, os, shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_HERE = Path(__file__).resolve().parent          # scripts/excel2/
_ROOT = _HERE.parent.parent                      # repo root
TEMPLATE_PATH = _ROOT / "bin" / "pqc_migration" / "BUKUKERJA_BENGKEL MIGRASI PQC 2025.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_json(path):
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return json.load(f)


def autosize(ws, max_width=80):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            # handle multiline strings
            max_len = max(max_len, max(len(ln) for ln in str(v).split("\n")))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def style_header_row(ws, row=4):
    """Style the header row (row 4 in template format) with dark blue fill."""
    fill  = PatternFill("solid", fgColor="1F4E79")
    font  = Font(color="FFFFFF", bold=True, size=12)
    align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for c in ws[row]:
        if c.value is not None:
            c.fill  = fill
            c.font  = font
            c.alignment = align
    ws.freeze_panes = f"A5"


def wrap_data_rows(ws, start_row=5):
    align = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=start_row):
        for c in row:
            c.alignment = align


def populate_sheet(ws, headers, rows, header_row=4):
    """
    Replace header + all data rows (from `header_row` down) with new content.
    Preserves rows 1–(header_row-1): title, description, blank preamble.
    """
    if ws.max_row >= header_row:
        ws.delete_rows(header_row, ws.max_row - header_row + 1)
    ws.append(headers)
    style_header_row(ws, row=header_row)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wrap_data_rows(ws, start_row=header_row + 1)
    autosize(ws)


# ---------------------------------------------------------------------------
# Severity helpers
# ---------------------------------------------------------------------------

_SEV_RANK = {
    "critical": 4, "high": 3, "medium": 2,
    "low": 1, "negligible": 0, "unknown": 0, "": 0, None: 0,
}


def worst_severity_rank(sev):
    return _SEV_RANK.get((sev or "").strip().lower(), 0)


def worst_severity_to_level(sev):
    s = (sev or "").strip().lower()
    if s == "critical": return "Sangat Tinggi"
    if s == "high":     return "Tinggi"
    if s == "medium":   return "Sederhana"
    if s in ("low", "negligible", "unknown"): return "Rendah"
    return ""


def score_to_risk_level(score: int) -> str:
    """Map risk score (impact × likelihood) to label matching the Risk Matrix."""
    if score >= 20: return "Risiko Sangat Tinggi"
    if score >= 15: return "Risiko Tinggi"
    if score >= 10: return "Risiko Sederhana"
    if score >= 5:  return "Risiko Rendah"
    return "Risiko Sangat Rendah"


# ---------------------------------------------------------------------------
# SBOM parsing
# ---------------------------------------------------------------------------

def parse_grype_worst_severity_by_pkg(grype):
    worst = {}
    for m in grype.get("matches") or []:
        art  = m.get("artifact") or {}
        vuln = m.get("vulnerability") or {}
        key  = (art.get("name") or "", art.get("version") or "", art.get("type") or "")
        sev  = vuln.get("severity") or ""
        if key not in worst or worst_severity_rank(sev) > worst_severity_rank(worst[key]):
            worst[key] = sev
    return worst


def extract_project_name(sbom):
    src    = sbom.get("source") or {}
    target = src.get("target") or ""
    if target:
        return os.path.basename(os.path.normpath(target)) or target
    return "UnknownProject"


# ---------------------------------------------------------------------------
# 0_Inventory — auto-populated from all scan sources
# ---------------------------------------------------------------------------

INVENTORY_HEADERS = [
    "#", "Asset Type", "Asset Name / Identifier", "Location / Owner",
    "Cryptographic Functionality Present?", "Examples of Algorithms Used",
    "SBOM/CBOM Available?", "Migration Readiness Level", "Notes / Action Items",
]

_BROKEN_ALGOS   = {"md5", "sha1", "sha-1", "sha-2", "des", "3des", "rc4", "rc2",
                    "blowfish", "seed", "idea", "rot13", "null"}
_QWEAK_ALGOS    = {"rsa", "ecdsa", "ecdh", "dsa", "dh", "ed25519", "x25519",
                    "aes-128", "aes128"}
_PQC_INDICATORS = {"ml-kem", "ml-dsa", "kyber", "dilithium", "falcon",
                    "sntrup", "mlkem", "sphincs", "frodo"}


def _readiness(algos_text: str) -> str:
    t = algos_text.lower()
    if any(x in t for x in _PQC_INDICATORS) and not any(x in t for x in _BROKEN_ALGOS):
        return "High"
    if any(x in t for x in _BROKEN_ALGOS):
        return "Very Low"
    if any(x in t for x in _QWEAK_ALGOS):
        return "Low"
    if algos_text.strip():
        return "Medium"
    return "Unknown"


def _mini_text(recs: list, module_id: int, kind: int = 1) -> str:
    """Extract info-type text lines from a specific mini-pqc module."""
    return "; ".join(
        r["Text"] for r in recs
        if r.get("ModuleID") == module_id and r.get("Kind") == kind
    )


def _mini_has_critical(recs: list, module_id: int) -> bool:
    return any(r.get("Severity", 0) >= 3 and r.get("ModuleID") == module_id
               for r in recs)


def parse_inventory_rows(
    sbom:           dict,
    grype_worst:    dict,
    cbom_rows:      list,
    mini_pqc:       dict | None = None,
    tls_data:       dict | None = None,
    ssh_data:       dict | None = None,
    cert_data:      dict | None = None,
    config_data:    dict | None = None,
    starttls_data:  dict | None = None,
    ike_data:       dict | None = None,
    rdp_data:       dict | None = None,
    smb_data:       dict | None = None,
    db_data:        dict | None = None,
    snmp_data:      dict | None = None,
    kerberos_data:  dict | None = None,
) -> list:
    """Build 0_Inventory rows from all available scan sources."""
    rows = []
    idx  = 0

    def add(asset_type, name, location, crypto_present, algos, sbom_cbom, readiness, notes):
        nonlocal idx
        idx += 1
        rows.append({
            "#":                                   idx,
            "Asset Type":                          asset_type,
            "Asset Name / Identifier":             name,
            "Location / Owner":                    location,
            "Cryptographic Functionality Present?": crypto_present,
            "Examples of Algorithms Used":         algos,
            "SBOM/CBOM Available?":                sbom_cbom,
            "Migration Readiness Level":           readiness,
            "Notes / Action Items":                notes,
        })

    # ------------------------------------------------------------------ #
    # 1. SBOM — scanned application / project                             #
    # ------------------------------------------------------------------ #
    project = extract_project_name(sbom)
    pkg_count = len(sbom.get("artifacts") or [])
    vuln_count = sum(1 for m in (grype_worst or {}).values() if m)
    algos_from_cbom = ", ".join(sorted({
        r.get("Algorithm Used", "") for r in cbom_rows
        if r.get("Algorithm Used")
    })[:6])
    add(
        asset_type  = "Application Stack",
        name        = project,
        location    = "Target Scan Path",
        crypto_present = "Yes" if cbom_rows else "No",
        algos       = algos_from_cbom or "(see CBOM sheet)",
        sbom_cbom   = "Yes (SBOM + CBOM)",
        readiness   = _readiness(algos_from_cbom),
        notes       = (
            f"{pkg_count} software component(s) in SBOM; "
            f"{vuln_count} with known CVEs; "
            f"{len(cbom_rows)} crypto usage finding(s) in CBOM."
        ),
    )

    # ------------------------------------------------------------------ #
    # 2. mini-pqc system findings (one row per detected component)         #
    # ------------------------------------------------------------------ #
    if mini_pqc:
        recs = mini_pqc.get("recommendations") or []

        # Operating System (Module 1)
        os_text = _mini_text(recs, module_id=1, kind=1)
        if os_text:
            add(
                asset_type  = "Operating System",
                name        = os_text.split(";")[0][:80],
                location    = mini_pqc.get("server_ip", "Scan Target"),
                crypto_present = "Yes",
                algos       = "Kernel crypto API, system libraries",
                sbom_cbom   = "Yes (mini-pqc)",
                readiness   = "Low" if _mini_has_critical(recs, 1) else "Medium",
                notes       = (
                    "Critical: Install OpenSSL 3.x for PQC provider support."
                    if _mini_has_critical(recs, 1) else
                    "Review mini-pqc recommendations for OS-level hardening."
                ),
            )

        # OpenSSL (Module 11 + 14)
        ssl_info = next(
            (r["Text"] for r in recs
             if r.get("ModuleID") in (11, 14) and "OpenSSL" in r.get("Text","")
             and r.get("Kind") == 1),
            None,
        )
        if ssl_info:
            pqc_native = any(
                "ML-KEM" in r.get("Text","") or "ML-DSA" in r.get("Text","")
                for r in recs if r.get("ModuleID") == 14
            )
            add(
                asset_type  = "Cryptographic Library",
                name        = ssl_info[:80],
                location    = mini_pqc.get("server_ip", "Scan Target"),
                crypto_present = "Yes",
                algos       = "RSA, AES, ECDSA, SHA-2" + (", ML-KEM, ML-DSA" if pqc_native else ""),
                sbom_cbom   = "Yes (mini-pqc)",
                readiness   = "High" if pqc_native else "Low",
                notes       = (
                    "OpenSSL 3.5+ with native ML-KEM/ML-DSA support detected. Verify OQS provider for additional PQC algorithms."
                    if pqc_native else
                    "Upgrade to OpenSSL 3.5+ for native ML-KEM (FIPS 203) and ML-DSA (FIPS 204) support."
                ),
            )

        # OpenSSH (Module 6)
        ssh_server = next(
            (r["Text"] for r in recs
             if r.get("ModuleID") == 6 and "Server" in r.get("Text","")
             and r.get("Kind") == 1),
            None,
        )
        if ssh_server and "Not installed" not in ssh_server:
            pqc_ssh = any(
                "hybrid" in r.get("Text","").lower() or "sntrup" in r.get("Text","").lower()
                for r in recs if r.get("ModuleID") == 6
            )
            add(
                asset_type  = "Network Service (SSH)",
                name        = ssh_server[:80],
                location    = mini_pqc.get("server_ip", "Scan Target"),
                crypto_present = "Yes",
                algos       = "ECDH, RSA, AES, ChaCha20" + (", sntrup761x25519" if pqc_ssh else ""),
                sbom_cbom   = "Yes (mini-pqc)",
                readiness   = "Medium" if pqc_ssh else "Low",
                notes       = (
                    "PQC-hybrid KEX detected in SSH. Verify sntrup761x25519 is preferred."
                    if pqc_ssh else
                    "Upgrade to OpenSSH 8.5+ and add sntrup761x25519-sha512 to KexAlgorithms."
                ),
            )

        # Web Server / Nginx (Module 5)
        nginx_info = next(
            (r["Text"] for r in recs
             if r.get("ModuleID") == 5 and r.get("Kind") == 1),
            None,
        )
        if nginx_info and "Not installed" not in nginx_info:
            add(
                asset_type  = "Web Server",
                name        = nginx_info[:80],
                location    = mini_pqc.get("server_ip", "Scan Target"),
                crypto_present = "Yes",
                algos       = "TLS (see 2_CBOM sheet)",
                sbom_cbom   = "Yes (mini-pqc)",
                readiness   = "Low",
                notes       = _mini_text(recs, module_id=5, kind=0)[:200] or
                              "Configure with OpenSSL 3.x + OQS provider for PQC TLS support.",
            )

        # Java (Module 10)
        java_info = next(
            (r["Text"] for r in recs
             if r.get("ModuleID") == 10 and "Java" in r.get("Text","")
             and r.get("Kind") == 1),
            None,
        )
        if java_info and "Not installed" not in java_info:
            add(
                asset_type  = "Runtime Environment",
                name        = java_info[:80],
                location    = mini_pqc.get("server_ip", "Scan Target"),
                crypto_present = "Yes",
                algos       = "RSA, AES, ECDSA (JCA/JCE)",
                sbom_cbom   = "Yes (mini-pqc)",
                readiness   = "Low",
                notes       = (
                    "Add BouncyCastle PQC provider (BCPQC) for ML-KEM/ML-DSA support. "
                    "Java 21+ includes experimental PQC in SunEC provider."
                ),
            )

    # ------------------------------------------------------------------ #
    # 3. Application components from CBOM (one row per distinct app)       #
    # ------------------------------------------------------------------ #
    app_algos: dict[str, set] = {}
    for r in cbom_rows:
        app = r.get("System / Application", "")
        algo = r.get("Algorithm Used", "")
        if app and algo:
            app_algos.setdefault(app, set()).add(algo)

    for app_name, algos_set in app_algos.items():
        algos_str = ", ".join(sorted(algos_set)[:6])
        add(
            asset_type  = "Application Code",
            name        = app_name,
            location    = "Target Scan Path",
            crypto_present = "Yes",
            algos       = algos_str,
            sbom_cbom   = "Yes (CBOM)",
            readiness   = _readiness(algos_str),
            notes       = (
                f"{len(algos_set)} distinct cryptographic algorithm(s) detected. "
                "See 2_CBOM sheet for file-level details and 4_RiskAssessment for mitigation."
            ),
        )

    # ------------------------------------------------------------------ #
    # 4. TLS network services                                              #
    # ------------------------------------------------------------------ #
    if tls_data:
        for chk in tls_data.get("tls_checks", []):
            if chk.get("error"):
                continue
            host = chk.get("hostname", "")
            port = chk.get("port", 443)
            cipher_d = chk.get("cipher_suite") or {}
            tls_ver  = chk.get("tls_version", "?")
            cipher   = cipher_d.get("name", "")
            pqc_ind  = chk.get("pqc_indicators", [])
            weak     = chk.get("weak_findings", [])
            algos_str = f"{tls_ver}, {cipher}" if cipher else tls_ver
            add(
                asset_type  = "Network Service (TLS)",
                name        = f"{host}:{port}",
                location    = host,
                crypto_present = "Yes",
                algos       = algos_str,
                sbom_cbom   = "Yes (TLS scan)",
                readiness   = "High" if pqc_ind else ("Low" if weak else "Medium"),
                notes       = (
                    f"PQC indicator: {', '.join(pqc_ind)}." if pqc_ind else
                    f"{len(weak)} weak finding(s). Migrate to ML-KEM hybrid KEM in TLS 1.3."
                ),
            )

    # ------------------------------------------------------------------ #
    # 5. SSH network services                                              #
    # ------------------------------------------------------------------ #
    if ssh_data:
        for chk in ssh_data.get("ssh_checks", []):
            if chk.get("error"):
                continue
            host = chk.get("hostname", "")
            port = chk.get("port", 22)
            banner = chk.get("banner", "")
            kex  = (chk.get("kex_init") or {}).get("kex_algorithms", [])
            pqc_ind = chk.get("pqc_indicators", [])
            weak    = chk.get("weak_findings", [])
            kex_str = ", ".join(kex[:4]) + ("..." if len(kex) > 4 else "")
            add(
                asset_type  = "Network Service (SSH)",
                name        = f"{host}:{port} ({banner})" if banner else f"{host}:{port}",
                location    = host,
                crypto_present = "Yes",
                algos       = kex_str,
                sbom_cbom   = "Yes (SSH scan)",
                readiness   = "High" if pqc_ind else ("Low" if weak else "Medium"),
                notes       = (
                    f"PQC-hybrid KEX: {', '.join(pqc_ind)}." if pqc_ind else
                    f"{len(weak)} weak algorithm(s). Add sntrup761x25519-sha512@openssh.com to KexAlgorithms."
                ),
            )

    # ------------------------------------------------------------------ #
    # 6. Certificate / PKI files                                           #
    # ------------------------------------------------------------------ #
    if cert_data:
        cert_entries = cert_data.get("cert_file_scan", [])
        total_certs  = len(cert_entries)
        total_findings = cert_data.get("summary", {}).get("total_findings", 0)
        if cert_entries:
            algos_found = ", ".join(sorted({
                e.get("algorithm", "") for e in cert_entries if e.get("algorithm")
            })[:6])
            expired  = sum(1 for e in cert_entries
                          for f in e.get("findings", []) if "Expired" in f.get("type",""))
            priv_keys = sum(1 for e in cert_entries if e.get("type") == "Private Key")
            add(
                asset_type  = "Certificate / PKI",
                name        = f"Certificate Files ({total_certs} found)",
                location    = "Target Scan Path",
                crypto_present = "Yes",
                algos       = algos_found or "(various)",
                sbom_cbom   = "Yes (cert scan)",
                readiness   = _readiness(algos_found),
                notes       = (
                    f"{total_certs} cert/key file(s) found; "
                    f"{expired} expired; "
                    f"{priv_keys} private key file(s) — consider HSM storage; "
                    f"{total_findings} total finding(s). See 2_CBOM sheet for details."
                ),
            )

    # ------------------------------------------------------------------ #
    # 7. Configuration files with crypto findings                          #
    # ------------------------------------------------------------------ #
    if config_data:
        config_entries = config_data.get("config_scan", [])
        total_findings = config_data.get("summary", {}).get("total_findings", 0)
        if config_entries:
            all_findings = [f for e in config_entries for f in e.get("findings",[])]
            critical_n = sum(1 for f in all_findings if f.get("severity") == "Critical")
            high_n     = sum(1 for f in all_findings if f.get("severity") == "High")
            types_str  = ", ".join(sorted({f.get("finding_type","") for f in all_findings[:10]}))
            add(
                asset_type  = "Configuration Files",
                name        = f"Server/App Config ({len(config_entries)} file(s) with findings)",
                location    = "Target Scan Path",
                crypto_present = "Yes",
                algos       = types_str[:80] or "(weak cipher/protocol strings found)",
                sbom_cbom   = "Yes (config scan)",
                readiness   = "Very Low" if critical_n > 0 else ("Low" if high_n > 0 else "Medium"),
                notes       = (
                    f"{total_findings} finding(s): "
                    f"{critical_n} Critical, {high_n} High. "
                    "See 2_CBOM sheet for file-level details."
                ),
            )

    # ------------------------------------------------------------------ #
    # 8. STARTTLS services                                                 #
    # ------------------------------------------------------------------ #
    if starttls_data:
        for chk in starttls_data.get("starttls_checks", []):
            if chk.get("error"):
                continue
            host    = chk.get("hostname", "")
            port    = chk.get("port", "")
            service = chk.get("service", "STARTTLS")
            cipher_d = chk.get("cipher_suite") or {}
            tls_ver  = chk.get("tls_version", "?")
            cipher   = cipher_d.get("name", "")
            pqc_ind  = chk.get("pqc_indicators", [])
            weak     = chk.get("weak_findings", [])
            supported = chk.get("starttls_supported", False)
            algos_str = f"{tls_ver}, {cipher}" if cipher else (tls_ver or "No TLS")
            add(
                asset_type  = f"Network Service ({service})",
                name        = f"{host}:{port} ({service})",
                location    = host,
                crypto_present = "Yes" if supported else "No",
                algos       = algos_str,
                sbom_cbom   = "Yes (STARTTLS scan)",
                readiness   = "High" if pqc_ind else ("Low" if weak else ("Very Low" if not supported else "Medium")),
                notes       = (
                    f"PQC indicator detected." if pqc_ind else
                    f"STARTTLS not supported — plaintext risk." if not supported else
                    f"{len(weak)} weak TLS finding(s)."
                ),
            )

    # ------------------------------------------------------------------ #
    # 9. IKE/IPsec VPN                                                    #
    # ------------------------------------------------------------------ #
    if ike_data:
        for chk in ike_data.get("ike_checks", []):
            if chk.get("error"):
                continue
            host    = chk.get("hostname", "")
            port    = chk.get("port", "")
            version = chk.get("ike_version", "IKE")
            dh      = (chk.get("preferred_dh_group") or {}).get("name", "Unknown DH")
            enc     = ", ".join(chk.get("encryption_algos", []))
            pqc_rdy = chk.get("pqc_ready", False)
            weak    = chk.get("weak_findings", [])
            ikev1   = chk.get("ikev1_enabled", False)
            add(
                asset_type  = "Network Service (VPN/IPsec)",
                name        = f"{host}:{port} ({version})",
                location    = host,
                crypto_present = "Yes",
                algos       = f"{dh}; {enc}" if enc else dh,
                sbom_cbom   = "Yes (IKE scan)",
                readiness   = "High" if pqc_rdy else ("Very Low" if ikev1 else "Low"),
                notes       = (
                    "PQC-hybrid DH detected." if pqc_rdy else
                    f"IKEv1 enabled (deprecated). {len(weak)} weak finding(s)." if ikev1 else
                    f"{len(weak)} weak finding(s). Migrate to IKEv2 with PQC hybrid DH."
                ),
            )

    # ------------------------------------------------------------------ #
    # 10. RDP                                                              #
    # ------------------------------------------------------------------ #
    if rdp_data:
        for chk in rdp_data.get("rdp_checks", []):
            if chk.get("error"):
                continue
            host    = chk.get("hostname", "")
            port    = chk.get("port", "")
            proto   = chk.get("security_protocol", "Unknown")
            cipher_d = chk.get("cipher_suite") or {}
            tls_ver  = chk.get("tls_version", "")
            cipher   = cipher_d.get("name", "")
            weak    = chk.get("weak_findings", [])
            classic = chk.get("classic_rdp_security", False)
            algos_str = f"{tls_ver}, {cipher}" if cipher else (proto or "Classic RDP")
            add(
                asset_type  = "Network Service (RDP)",
                name        = f"{host}:{port}",
                location    = host,
                crypto_present = "Yes",
                algos       = algos_str,
                sbom_cbom   = "Yes (RDP scan)",
                readiness   = "Very Low" if classic else ("Low" if weak else "Medium"),
                notes       = (
                    "Classic RDP security (no NLA/TLS) — RC4 encryption, no server auth." if classic else
                    f"{len(weak)} weak TLS cipher(s). Enable NLA and TLS 1.3."
                ),
            )

    # ------------------------------------------------------------------ #
    # 11. SMB                                                              #
    # ------------------------------------------------------------------ #
    if smb_data:
        for chk in smb_data.get("smb_checks", []):
            if chk.get("error"):
                continue
            host    = chk.get("hostname", "")
            port    = chk.get("port", "")
            dialect = chk.get("dialect", "SMB")
            enc_sup = chk.get("encryption_supported", False)
            smb1    = chk.get("smb1_enabled", False)
            ciphers = ", ".join(chk.get("cipher_ids", []))
            weak    = chk.get("weak_findings", [])
            add(
                asset_type  = "Network Service (SMB)",
                name        = f"{host}:{port} ({dialect})",
                location    = host,
                crypto_present = "Yes" if enc_sup else "Partial",
                algos       = ciphers or ("None (SMBv1)" if smb1 else "NTLM/Kerberos"),
                sbom_cbom   = "Yes (SMB scan)",
                readiness   = "Very Low" if smb1 else ("Low" if not enc_sup else "Medium"),
                notes       = (
                    "SMBv1 enabled — EternalBlue/WannaCry risk; no message encryption." if smb1 else
                    f"SMB encryption {'supported' if enc_sup else 'NOT supported'}. {len(weak)} weak finding(s)."
                ),
            )

    # ------------------------------------------------------------------ #
    # 12. Database services                                                #
    # ------------------------------------------------------------------ #
    if db_data:
        for chk in db_data.get("db_checks", []):
            if chk.get("error"):
                continue
            host    = chk.get("hostname", "")
            port    = chk.get("port", "")
            service = chk.get("service", "Database")
            tls_sup = chk.get("tls_supported", False)
            tls_ver = chk.get("tls_version", "")
            cipher_d = chk.get("cipher_suite") or {}
            cipher  = cipher_d.get("name", "")
            weak    = chk.get("weak_findings", [])
            algos_str = f"{tls_ver}, {cipher}" if cipher else ("No TLS" if not tls_sup else "TLS (cipher unknown)")
            add(
                asset_type  = f"Database Service ({service})",
                name        = f"{host}:{port} ({service})",
                location    = host,
                crypto_present = "Yes" if tls_sup else "No",
                algos       = algos_str,
                sbom_cbom   = "Yes (DB scan)",
                readiness   = "Very Low" if not tls_sup else ("Low" if weak else "Medium"),
                notes       = (
                    f"{service} does not support TLS — data in transit is unencrypted." if not tls_sup else
                    f"{len(weak)} weak TLS finding(s). Enforce TLS 1.3 with strong ciphers."
                ),
            )

    # ------------------------------------------------------------------ #
    # 13. SNMP agents                                                      #
    # ------------------------------------------------------------------ #
    if snmp_data:
        for chk in snmp_data.get("snmp_checks", []):
            if chk.get("error"):
                continue
            host = chk.get("hostname", "")
            port = chk.get("port", "")
            v1   = chk.get("v1_enabled", False)
            v2c  = chk.get("v2c_enabled", False)
            v3   = chk.get("v3_enabled", False)
            weak = chk.get("weak_findings", [])
            versions = ", ".join(filter(None, [
                "SNMPv1" if v1 else "",
                "SNMPv2c" if v2c else "",
                "SNMPv3" if v3 else "",
            ]))
            add(
                asset_type  = "Network Service (SNMP)",
                name        = f"{host}:{port}",
                location    = host,
                crypto_present = "Partial" if v3 else "No",
                algos       = "Community String (no crypto)" if (v1 or v2c) else "USM (MD5/SHA)",
                sbom_cbom   = "Yes (SNMP scan)",
                readiness   = "Very Low" if (v1 or v2c) else "Low",
                notes       = (
                    f"Versions detected: {versions}. "
                    + ("SNMPv1/v2c use cleartext community strings — no encryption. " if (v1 or v2c) else "")
                    + ("Disable v1/v2c; enforce SNMPv3 with AES-256 priv." if (v1 or v2c) else
                       "SNMPv3 only — verify auth/priv protocols.")
                ),
            )

    # ------------------------------------------------------------------ #
    # 14. Kerberos KDCs                                                    #
    # ------------------------------------------------------------------ #
    if kerberos_data:
        for chk in kerberos_data.get("kerberos_checks", []):
            if chk.get("error"):
                continue
            host   = chk.get("hostname", "")
            port   = chk.get("port", "")
            etypes = ", ".join(chk.get("etype_names", []))
            rc4    = chk.get("rc4_enabled", False)
            des    = chk.get("des_enabled", False)
            weak   = chk.get("weak_findings", [])
            add(
                asset_type  = "Authentication Service (Kerberos)",
                name        = f"{host}:{port} (KDC)",
                location    = host,
                crypto_present = "Yes",
                algos       = etypes or "Unknown etypes",
                sbom_cbom   = "Yes (Kerberos scan)",
                readiness   = "Very Low" if (rc4 or des) else ("Low" if weak else "Medium"),
                notes       = (
                    f"Encryption types: {etypes}. "
                    + ("RC4/DES enabled — plaintext-equivalent risk (Pass-the-Hash). " if (rc4 or des) else "")
                    + ("Disable RC4-HMAC and DES; enforce AES256-CTS-HMAC-SHA384-192." if (rc4 or des) else
                       "Verify AES256-CTS is the preferred etype.")
                ),
            )

    return rows


def extract_url_from_syft_artifact(art):
    md = art.get("metadata") or {}
    if isinstance(md, dict):
        for k in ("homepage", "url"):
            if md.get(k):
                return md[k]
        src = md.get("source")
        if isinstance(src, dict) and src.get("url"):
            return src["url"]
        dist = md.get("dist")
        if isinstance(dist, dict) and dist.get("url"):
            return dist["url"]
    return ""


def parse_sbom_rows(sbom, grype_worst):
    project = extract_project_name(sbom)
    rows = []
    for i, art in enumerate(sbom.get("artifacts") or [], start=1):
        name    = art.get("name") or ""
        version = art.get("version") or ""
        typ     = art.get("type") or ""
        url     = extract_url_from_syft_artifact(art)
        worst   = grype_worst.get((name, version, typ), "")
        rows.append({
            "#":                                          i,
            "System / Application":                       project,
            "Purpose / Usage":                            "",
            "URL":                                        url,
            "Services Mode":                              "",
            "Target Customer":                            "",
            "Software Component":                         f"{name} {version}".strip(),
            "Third-party Modules":                        "",
            "External APIs or Services":                  "",
            "Critical Level":                             worst_severity_to_level(worst),
            "Data Category":                              "",
            "Is the application/system currently in use?":"",
            "Application/System Developer":               "",
            "Vendor's Name":                              "",
            "Does the agency have expertise?":            "",
            "Does the agency have a special budget allocation?": "",
            "Link to CBOM":                               f"CBOM ({project})",
        })
    return rows


# ---------------------------------------------------------------------------
# CBOM parsing — deduplicated, with file path + line number
# ---------------------------------------------------------------------------

def _get_path(r: dict) -> str:
    path = r.get("path") or ""
    if not path:
        locs = r.get("locations") or []
        if locs:
            al   = ((locs[0].get("physicalLocation") or {}).get("artifactLocation") or {})
            path = al.get("uri") or ""
    return path


def _pick(meta: dict, *keys, default="") -> str:
    for k in keys:
        if k in meta and meta[k] not in (None, "", {}):
            return str(meta[k])
    return default


def _shorten_path(path: str) -> str:
    """Return the last 3 path segments so it fits in a cell."""
    if not path:
        return ""
    parts = Path(path).parts
    return str(Path(*parts[-3:])) if len(parts) >= 3 else path


def _extract_app_name(path: str) -> str:
    """
    Infer the application/module name from the file path.
    Takes the 3rd-from-last directory segment (project root heuristic).
    e.g. .../testing/DVWA/dvwa/includes/file.php → DVWA
    """
    if not path:
        return "(Unknown)"
    parts = Path(path).parts
    # Walk upward looking for the first "meaningful" directory name
    # (skip common generic names like 'includes', 'src', 'lib', 'app')
    _SKIP = {"includes", "src", "lib", "app", "core", "modules", "classes",
              "vendor", "node_modules", "dist", "build", "assets", "static",
              "public", "private", "common", "util", "utils", "helper", "helpers"}
    for seg in reversed(parts[:-1]):
        if seg and seg not in _SKIP and not seg.startswith("."):
            return seg
    return parts[-2] if len(parts) >= 2 else "(Unknown)"


def parse_cbom_rows(cbom: dict) -> list:
    """
    Parse Semgrep CBOM JSON into deduplicated rows.
    Deduplication key: (file_path, line_number, algorithm).
    When multiple rules fire on the same location, the best mitigation,
    PQC risk, and vulnerability description are merged.
    """
    raw_results = []
    if isinstance(cbom, dict):
        if isinstance(cbom.get("results"), list):
            raw_results = cbom["results"]
        elif isinstance(cbom.get("runs"), list):
            raw_results = (cbom["runs"][0].get("results") or []) if cbom["runs"] else []

    # --- deduplicate by (path, line, algo) ---
    merged: dict = {}   # key → dict of best fields

    for r in raw_results:
        meta     = ((r.get("extra") or {}).get("metadata") or {})
        cbom_m   = meta.get("cbom") or {}
        path     = _get_path(r)
        line     = (r.get("start") or {}).get("line", "")
        algo     = _pick(cbom_m, "algorithm_used") or _pick(meta, "algorithm_name", "algorithm_used", "algorithm")
        func     = _pick(cbom_m, "cryptographic_function") or _pick(meta, "cryptographic_function", "crypto_function")
        lib      = _pick(cbom_m, "library_module") or _pick(meta, "library_module", "library", "module")
        key_len  = _pick(cbom_m, "key_length") or _pick(meta, "key_length", "keylen")
        purpose  = _pick(cbom_m, "purpose_usage") or _pick(meta, "purpose_usage", "purpose", "usage")
        agility  = _pick(meta, "crypto_agility_support", "crypto_agility", "agility")
        pqc_risk = _pick(meta, "pqc_risk")
        mitigation = _pick(meta, "recommended_mitigation")
        vuln     = _pick(meta, "vulnerability")
        message  = _pick(r.get("extra") or {}, "message")

        dedup_key = (path, str(line), (algo or "").lower())

        if dedup_key not in merged:
            merged[dedup_key] = {
                "path":       path,
                "line":       line,
                "algo":       algo,
                "func":       func,
                "lib":        lib,
                "key_len":    key_len,
                "purpose":    purpose,
                "agility":    agility,
                "pqc_risk":   "",
                "mitigation": "",
                "vuln":       "",
                "message":    "",
            }

        slot = merged[dedup_key]
        # Always prefer more specific data over empty
        if pqc_risk   and not slot["pqc_risk"]:   slot["pqc_risk"]   = pqc_risk
        if mitigation  and not slot["mitigation"]:  slot["mitigation"] = mitigation
        if vuln        and not slot["vuln"]:        slot["vuln"]       = vuln
        if message     and not slot["message"]:     slot["message"]    = message
        if func        and not slot["func"]:        slot["func"]       = func
        if lib         and not slot["lib"]:         slot["lib"]        = lib
        if purpose     and not slot["purpose"]:     slot["purpose"]    = purpose
        if agility     and not slot["agility"]:     slot["agility"]    = agility

    # --- build output rows ---
    rows = []
    for idx, slot in enumerate(merged.values(), start=1):
        app_name   = _extract_app_name(slot["path"])
        short_path = _shorten_path(slot["path"])
        finding    = slot["vuln"] or (slot["message"][:120] if slot["message"] else "")

        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   app_name,
            "File Path":              short_path,
            "Line":                   slot["line"],
            "Cryptographic Function": slot["func"],
            "Algorithm Used":         slot["algo"],
            "Library / Module":       slot["lib"],
            "Key Length":             slot["key_len"],
            "Purpose / Usage":        slot["purpose"],
            "PQC Risk":               slot["pqc_risk"],
            "Crypto-Agility Support": slot["agility"],
            "Finding / Vulnerability": finding,
            # private — used by risk generator, not written to sheet
            "_path":       slot["path"],
            "_line":       slot["line"],
            "_mitigation": slot["mitigation"],
        })
    return rows


# ---------------------------------------------------------------------------
# Risk Register & Risk Assessment — algorithm-specific, uses Semgrep metadata
# ---------------------------------------------------------------------------

def _classify_single(a: str, f: str) -> str:
    """Classify a single (already lowercased) algorithm token."""
    if any(x in a for x in [
        "rsa", "ecdsa", "ecdh", "dsa", " dh", "dh-", "dh group", "dh (", "modp-", "modp ",
        "ecp-", "ecp ", "ed25519", "x25519", "x448", "curve25519", "curve448",
        "kyber", "dilithium", "falcon", "ml-kem", "ml-dsa", "ikev1", "ikev2",
    ]):
        return "asym"
    if any(x in a for x in ["md5", "sha-1", "sha1", "hmac-md5", "hmac-sha1",
                              "usm (md5", "usm (sha-1", "usm (sha1"]):
        return "hash_weak"
    if any(x in a for x in ["sha-256", "sha256", "sha-384", "sha-512", "sha3",
                              "blake", "hmac-sha256", "hmac-sha384", "hmac-sha512"]):
        return "hash_strong"
    if any(x in a for x in ["des", "3des", "rc4", "rc2", "blowfish", "seed",
                              "idea", "ecb", "arcfour", "rc4-hmac", "des-cbc",
                              "des-md5", "des-sha1"]):
        return "sym_weak"
    if any(x in a for x in ["aes", "chacha20", "chacha", "aes128-cts", "aes256-cts"]):
        return "sym_strong"
    if any(x in a for x in ["community string", "snmpv1", "snmpv2c"]):
        return "obfuscation"
    if any(x in a for x in ["rand", "mt_rand", "uniqid", "math.random"]) or "rng" in f or "random" in f:
        return "rng"
    if any(x in a for x in ["base64", "rot13", "hex"]):
        return "obfuscation"
    return "other"


# Severity ordering: lower index = worse/higher priority classification
_CLASS_ORDER = ["asym", "sym_weak", "hash_weak", "obfuscation", "rng",
                "hash_strong", "sym_strong", "other"]


def _classify_algo(algo: str, func: str) -> str:
    """Return one of: asym | hash_weak | hash_strong | sym_weak | sym_strong | rng | obfuscation | other

    Handles comma-separated algorithm lists (e.g. Kerberos mixed-etype strings like
    "ArcFour-HMAC, AES256-CTS-HMAC-SHA1-96") by classifying each token and returning
    the worst (highest-severity) classification.
    """
    f = func.lower()
    tokens = [t.strip().lower() for t in algo.split(",") if t.strip()]
    if not tokens:
        return "other"
    results = [_classify_single(t, f) for t in tokens]
    # Return the result with the lowest index in _CLASS_ORDER (= worst severity)
    return min(results, key=lambda c: _CLASS_ORDER.index(c) if c in _CLASS_ORDER else len(_CLASS_ORDER))


def _risk_for_algo(kind: str, algo: str, func: str, purpose: str, file_base: str, line) -> tuple:
    """
    Returns (risiko_rr, risiko_ra, punca_risiko, tahap_kritikal, impak, likelihood, kawalan)
    All strings are actionable and specific to the algorithm class.
    """
    loc = f"{file_base}:{line}" if line else file_base

    if kind == "asym":
        risiko_rr = (
            f"Kelemahan PQC: Algoritma asimetri {algo} terdedah kepada Algoritma Shor — "
            f"kunci dan tandatangan boleh dipecahkan oleh komputer kuantum (CRQC)"
        )
        risiko_ra = "Pendedahan kepada Algoritma Shor (Quantum Integer Factorization / Discrete Logarithm)"
        punca = (
            f"Algoritma {algo} digunakan untuk '{purpose or func}' dalam {loc}. "
            f"Semua skema berasaskan masalah faktorisasi integer dan logaritma diskret "
            f"(RSA, ECC, DH, DSA) akan dipecahkan dalam masa O(log³N) oleh Algoritma Shor "
            f"pada CRQC dengan ~ 4096 qubit logik yang stabil."
        )
        tahap   = "Kritikal"
        impak   = 5
        like    = 5
        kawalan = (
            "Infrastruktur PKI X.509 sedia ada; pematuhan TLS 1.3; "
            "Dasar keselamatan siber organisasi. "
            "TIADA kawalan yang mencukupi untuk ancaman kuantum."
        )

    elif kind == "hash_weak":
        risiko_rr = (
            f"Kelemahan Kriptografi: Fungsi hash {algo} telah dipecahkan — "
            f"serangan collision dan pra-imej boleh dilakukan"
        )
        risiko_ra = "Pendedahan kepada serangan collision (MD5 sejak 2004, SHA-1: SHAttered 2017)"
        punca = (
            f"{algo} digunakan untuk '{purpose or func}' dalam {loc}. "
            f"MD5 collision boleh dijana dalam < 1 saat pada perkakasan biasa. "
            f"SHA-1 collision pertama dibuktikan oleh Google (SHAttered, 2017). "
            f"Kedua-dua tidak selamat untuk mana-mana tujuan keselamatan."
        )
        tahap   = "Kritikal"
        impak   = 4
        like    = 4
        kawalan = (
            "Semakan kod statik (SAST) standard; dasar penggunaan hash organisasi. "
            "Kawalan tidak mencukupi — MD5/SHA-1 masih digunakan dalam kod aktif."
        )

    elif kind == "hash_strong":
        risiko_rr = (
            f"Pertimbangan PQC: Fungsi hash {algo} selamat secara klasik tetapi "
            f"kekuatan berkesan dikurangkan 50% oleh Algoritma Grover"
        )
        risiko_ra = "Pendedahan separa kepada Algoritma Grover (Quantum Search)"
        punca = (
            f"{algo} digunakan untuk '{purpose or func}' dalam {loc}. "
            f"Algoritma Grover mengurangkan kekuatan berkesan dari N bit kepada N/2 bit. "
            f"SHA-256 (256-bit) → berkesan 128-bit di era kuantum. Masih selamat jika saiz output mencukupi."
        )
        tahap   = "Sederhana"
        impak   = 2
        like    = 2
        kawalan = (
            f"Penggunaan {algo} adalah pematuhan NIST semasa. "
            "Pantauan perkembangan piawaian PQC NIST berterusan."
        )

    elif kind == "sym_weak":
        risiko_rr = (
            f"Kelemahan Kriptografi: Sifer simetri usang {algo} — "
            f"terdedah kepada serangan penyahsulitan dan serangan ke atas blok pendek"
        )
        risiko_ra = f"Sifer {algo} lemah/usang — serangan brute-force dan known-plaintext boleh dilakukan"
        punca = (
            f"{algo} digunakan untuk '{purpose or func}' dalam {loc}. "
            f"DES/3DES terdedah kepada Sweet32 (CVE-2016-2183, blok 64-bit). "
            f"RC4 telah dipecahkan sepenuhnya (RFC 7465). "
            f"Semua sifer warisan ini telah dialih keluar dari TLS 1.3."
        )
        tahap   = "Kritikal"
        impak   = 4
        like    = 3
        kawalan = (
            "Dasar TLS organisasi (mungkin melarang RC4). "
            "Namun kawalan tidak dipatuhi dalam kod aplikasi — sifer lemah masih digunakan."
        )

    elif kind == "sym_strong":
        risiko_rr = (
            f"Pertimbangan PQC: Sifer simetri {algo} selamat secara klasik; "
            f"kekuatan berkesan dikurangkan 50% oleh Algoritma Grover"
        )
        risiko_ra = "Pendedahan separa kepada Algoritma Grover (Quantum Search)"
        punca = (
            f"{algo} digunakan untuk '{purpose or func}' dalam {loc}. "
            f"AES-128 berkesan menjadi ~64-bit di era kuantum (tidak mencukupi). "
            f"AES-256 berkesan menjadi ~128-bit — masih selamat menurut NIST SP 800-227."
        )
        tahap   = "Rendah"
        impak   = 2
        like    = 2
        kawalan = (
            f"Penggunaan {algo} adalah pematuhan NIST semasa. "
            "Pastikan mod AEAD (GCM) digunakan dan saiz kunci ≥ 256-bit."
        )

    elif kind == "rng":
        risiko_rr = (
            f"Kelemahan Keselamatan: Penjana nombor rawak tidak selamat ({algo}) "
            f"boleh menjejaskan kerahsiaan kunci, token, dan sesi"
        )
        risiko_ra = "Penjana nombor rawak boleh diramal — kunci dan token kripto boleh diterka"
        punca = (
            f"{algo} digunakan untuk '{purpose or func}' dalam {loc}. "
            f"Fungsi seperti rand(), mt_rand(), Math.random(), dan uniqid() BUKAN CSPRNG. "
            f"Entropi rendah membolehkan penyerang menjangka nilai yang dijana."
        )
        tahap   = "Tinggi"
        impak   = 4
        like    = 3
        kawalan = (
            "Semakan kod statik (SAST) standard. "
            "Kawalan tidak mencukupi — CSPRNG tidak dikuatkuasakan dalam kod."
        )

    elif kind == "obfuscation":
        risiko_rr = (
            f"Kelemahan Reka Bentuk: {algo} digunakan sebagai 'penyulitan' "
            f"tetapi tidak memberikan kerahsiaan kriptografi"
        )
        risiko_ra = f"Salah faham kriptografi — {algo} bukan penyulitan, data masih boleh dihuraikan"
        punca = (
            f"{algo} digunakan sebagai pengganti penyulitan dalam {loc}. "
            f"Base64 adalah pengekodan (encoding) bukan penyulitan — sesiapa boleh nyahkod. "
            f"ROT13 adalah penggantian mudah — tiada keselamatan kriptografi."
        )
        tahap   = "Tinggi"
        impak   = 3
        like    = 4
        kawalan = (
            "Tiada kawalan kriptografi yang sedia ada bagi kes ini. "
            "Data mungkin kelihatan 'tersembunyi' tetapi tidak dilindungi."
        )

    else:
        risiko_rr = f"Penggunaan kriptografi ({algo}) memerlukan semakan dan pengkelasan PQC"
        risiko_ra = "Pendedahan PQC tidak dapat ditentukan — pengkelasan lanjut diperlukan"
        punca = (
            f"Penggunaan {algo} ({func}) dalam {loc} tidak dapat dikelaskan secara automatik. "
            f"Semakan manual diperlukan untuk menentukan risiko PQC."
        )
        tahap   = "Sederhana"
        impak   = 3
        like    = 2
        kawalan = "Semakan kod manual dan inventori kriptografi organisasi."

    return risiko_rr, risiko_ra, punca, tahap, impak, like, kawalan


def _mitigation_plan(kind: str, algo: str, semgrep_mitigation: str) -> str:
    """
    Return a specific mitigation plan.
    Prefers Semgrep's own recommended_mitigation; enriches it with NIST references.
    """
    nist_suffix = {
        "asym": (
            "\nMigrasi kepada:\n"
            "  • ML-KEM (CRYSTALS-Kyber) — NIST FIPS 203: pertukaran kunci / KEM\n"
            "  • ML-DSA (CRYSTALS-Dilithium) — NIST FIPS 204: tandatangan digital\n"
            "  • SLH-DSA (SPHINCS+) — NIST FIPS 205: tandatangan berasaskan hash\n"
            "Langkah interim: gunakan RSA-3072 / ECDSA P-384 minimum sambil menunggu migrasi penuh."
        ),
        "hash_weak": (
            "\nGantikan dengan:\n"
            "  • SHA-256 atau SHA-3-256 (minimum)\n"
            "  • SHA-512 atau SHA-3-512 untuk aplikasi kritikal\n"
            "  • Untuk hash kata laluan: gunakan Argon2id (OWASP recommended)"
        ),
        "sym_weak": (
            "\nGantikan dengan:\n"
            "  • AES-256-GCM (AEAD — disyorkan)\n"
            "  • ChaCha20-Poly1305 (alternatif AEAD)\n"
            "  • Jangan sekali-kali gunakan mod ECB"
        ),
        "rng": (
            "\nGantikan dengan:\n"
            "  • PHP: random_bytes() / random_int()\n"
            "  • Python: secrets.token_bytes() / os.urandom()\n"
            "  • JavaScript: crypto.getRandomValues() / crypto.randomBytes()\n"
            "  • Java: SecureRandom"
        ),
        "obfuscation": (
            "\nJika kerahsiaan diperlukan, gunakan penyulitan sebenar:\n"
            "  • AES-256-GCM (simetri)\n"
            "  • Atau gunakan pustaka kriptografi seperti libsodium / NaCl"
        ),
    }

    base = semgrep_mitigation.strip() if semgrep_mitigation else ""
    suffix = nist_suffix.get(kind, "")

    if base and suffix:
        return base + suffix
    elif base:
        return base
    elif suffix:
        return suffix.strip()
    else:
        return "Lakukan inventori kriptografi dan ikuti panduan migrasi PQC NIST SP 800-227."


def risk_rows_from_cbom(cbom_rows: list) -> tuple:
    rr_rows, ra_rows = [], []

    for i, row in enumerate(cbom_rows, start=1):
        algo    = row.get("Algorithm Used", "") or ""
        func    = row.get("Cryptographic Function", "") or ""
        purpose = row.get("Purpose / Usage", "") or ""
        app     = row.get("System / Application", "")
        file_path = row.get("_path", row.get("File Path", ""))
        line    = row.get("_line", row.get("Line", ""))
        semgrep_mit = row.get("_mitigation", "")

        file_base = os.path.basename(file_path) if file_path else app
        kind      = _classify_algo(algo, func)

        risiko_rr, risiko_ra, punca, tahap, impak, like, kawalan = \
            _risk_for_algo(kind, algo, func, purpose, file_base, line)

        mitigation = _mitigation_plan(kind, algo, semgrep_mit)
        score      = impak * like

        rr_rows.append({
            "#":                                        i,
            "Nama Sistem/ Perkakasan/Perisian":         app,
            "Jenis Aset\nAplikasi/Perkakasan/Perisian": "Application Code",
            "Algoritma Kriptografi":                    algo,
            "Kegunaan Algoritma Kriptografi":           purpose or func,
            "Tahap Kritikal":                           tahap,
            "Risiko":                                   risiko_rr,
            "Pemilik Risiko":                           "IT Security Team / Application Owner",
        })
        ra_rows.append({
            "#":                                   i,
            "Nama Sistem/ Perkakasan/Perisian":    app,
            "Algoritma Kriptografi":               algo,
            "Risiko":                              risiko_ra,
            "Punca Risiko":                        punca,
            "Impak":                               impak,
            "Kemungkinan\n(Likelihood)":           like,
            "Skor Risiko":                         score,
            "Risk Level":                          score_to_risk_level(score),
            "Kawalan Sedia Ada":                   kawalan,
            "Mitigation Plan":                     mitigation,
        })

    return rr_rows, ra_rows


# ---------------------------------------------------------------------------
# Sheet headers (must match column order in template row 4)
# ---------------------------------------------------------------------------

SBOM_HEADERS = [
    "#", "System / Application", "Purpose / Usage", "URL", "Services Mode",
    "Target Customer", "Software Component", "Third-party Modules",
    "External APIs or Services", "Critical Level", "Data Category",
    "Is the application/system currently in use?", "Application/System Developer",
    "Vendor's Name", "Does the agency have expertise?",
    "Does the agency have a special budget allocation?", "Link to CBOM",
]

CBOM_HEADERS = [
    "# (CBOM)", "System / Application", "File Path", "Line",
    "Cryptographic Function", "Algorithm Used", "Library / Module",
    "Key Length", "Purpose / Usage", "PQC Risk",
    "Crypto-Agility Support", "Finding / Vulnerability",
]

RR_HEADERS = [
    "#", "Nama Sistem/ Perkakasan/Perisian",
    "Jenis Aset\nAplikasi/Perkakasan/Perisian",
    "Algoritma Kriptografi", "Kegunaan Algoritma Kriptografi",
    "Tahap Kritikal", "Risiko", "Pemilik Risiko",
]

RA_HEADERS = [
    "#", "Nama Sistem/ Perkakasan/Perisian", "Algoritma Kriptografi",
    "Risiko", "Punca Risiko", "Impak", "Kemungkinan\n(Likelihood)",
    "Skor Risiko", "Risk Level", "Kawalan Sedia Ada", "Mitigation Plan",
]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Convert scan data (TLS/SSH/Cert/Config) → CBOM-compatible rows for sheets 0-4
# ---------------------------------------------------------------------------

def _pqc_risk_from_sev(severity: str) -> str:
    s = (severity or "").strip().lower()
    if s in ("critical", "high"):
        return "High"
    if s == "medium":
        return "Medium"
    return "Low"


def _extract_algo_from_cipher(cipher_suite: str) -> str:
    """Extract the primary key-exchange/auth algorithm from a TLS cipher suite name."""
    if not cipher_suite:
        return ""
    cs = cipher_suite.upper()
    if "ECDHE" in cs or "ECDH_" in cs:
        return "ECDH"
    if "DHE_" in cs or "_DH_" in cs:
        return "DH"
    if "RSA" in cs:
        return "RSA"
    if "3DES" in cs:
        return "3DES"
    if "RC4" in cs:
        return "RC4"
    if "AES_128" in cs:
        return "AES-128"
    if "AES_256" in cs:
        return "AES-256"
    if "CHACHA20" in cs:
        return "ChaCha20"
    return cipher_suite


def tls_to_cbom_rows(tls_data: dict, start_idx: int = 1) -> list:
    """Convert TLS scan results (sheet 5) into CBOM-compatible rows."""
    rows = []
    idx = start_idx
    sev_rank = {"critical": 4, "high": 3, "medium": 2, "low": 1}
    for chk in tls_data.get("tls_checks", []):
        if chk.get("error"):
            continue
        host   = chk.get("hostname", "")
        port   = chk.get("port", "")
        system = f"{host}:{port}" if port else host
        cipher_d   = chk.get("cipher_suite") or {}
        tls_ver    = chk.get("tls_version", "")
        pqc_ind    = chk.get("pqc_indicators", [])
        weak       = chk.get("weak_findings", [])
        if not weak:
            continue
        recs       = "\n".join(chk.get("recommendations", []))
        cipher_name = cipher_d.get("name", "")
        algo = _extract_algo_from_cipher(cipher_name) or cipher_name
        weak_str = "; ".join(
            f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}"
            for w in weak
        )
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": "TLS Key Exchange / Authentication",
            "Algorithm Used":         algo,
            "Library / Module":       f"TLS {tls_ver}".strip(),
            "Key Length":             str(cipher_d.get("key_bits", "")),
            "Purpose / Usage":        "TLS/SSL encrypted connection",
            "PQC Risk":               "Low" if pqc_ind else ("High" if weak else "Medium"),
            "Crypto-Agility Support": "Partial",
            "Finding / Vulnerability": weak_str or "No weak findings detected",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


def ssh_to_cbom_rows(ssh_data: dict, start_idx: int = 1) -> list:
    """Convert SSH scan results (sheet 6) into CBOM-compatible rows."""
    rows = []
    idx = start_idx
    for chk in ssh_data.get("ssh_checks", []):
        if chk.get("error"):
            continue
        host   = chk.get("hostname", "")
        port   = chk.get("port", "")
        system = f"{host}:{port}" if port else host
        banner = chk.get("banner", "")
        weak    = chk.get("weak_findings", [])
        if not weak:
            continue
        recs    = "\n".join(chk.get("recommendations", []))
        lib_ver = banner[:60] if banner else "SSH"
        for wf in weak:
            rows.append({
                "# (CBOM)":               f"CBOM #{idx}",
                "System / Application":   system,
                "File Path":              "",
                "Line":                   "",
                "Cryptographic Function": wf.get("category", "SSH Algorithm"),
                "Algorithm Used":         wf.get("algorithm", ""),
                "Library / Module":       lib_ver,
                "Key Length":             "",
                "Purpose / Usage":        "SSH remote access",
                "PQC Risk":               _pqc_risk_from_sev(wf.get("severity", "")),
                "Crypto-Agility Support": "No",
                "Finding / Vulnerability": wf.get("detail", ""),
                "_path": system, "_line": "", "_mitigation": wf.get("recommendation", ""),
            })
            idx += 1
    return rows


def cert_to_cbom_rows(cert_data: dict, start_idx: int = 1) -> list:
    """Convert certificate/key file scan results (sheet 7) into CBOM-compatible rows."""
    rows = []
    idx = start_idx
    for entry in cert_data.get("cert_file_scan", []):
        file_path = entry.get("file", "")
        algo      = entry.get("algorithm", "") or entry.get("signature_algorithm", "")
        key_bits  = entry.get("key_bits", "")
        qrisk     = entry.get("quantum_risk", "")
        cert_type = entry.get("type", "")
        subject   = entry.get("subject_cn", "")
        for f in entry.get("findings", []):
            rows.append({
                "# (CBOM)":               f"CBOM #{idx}",
                "System / Application":   subject or os.path.basename(file_path),
                "File Path":              file_path,
                "Line":                   "",
                "Cryptographic Function": cert_type or "Digital Certificate",
                "Algorithm Used":         algo,
                "Library / Module":       "X.509 / PKI",
                "Key Length":             str(key_bits) if key_bits else "",
                "Purpose / Usage":        "Certificate / Key File",
                "PQC Risk":               qrisk or _pqc_risk_from_sev(f.get("severity", "")),
                "Crypto-Agility Support": "No",
                "Finding / Vulnerability": f.get("detail", ""),
                "_path": file_path, "_line": "", "_mitigation": f.get("recommendation", ""),
            })
            idx += 1
    return rows


def config_to_cbom_rows(config_data: dict, start_idx: int = 1) -> list:
    """Convert config file weak-crypto scan results (sheet 8) into CBOM-compatible rows."""
    rows = []
    idx = start_idx
    for entry in config_data.get("config_scan", []):
        file_path = entry.get("file", "")
        for f in entry.get("findings", []):
            rows.append({
                "# (CBOM)":               f"CBOM #{idx}",
                "System / Application":   os.path.basename(file_path),
                "File Path":              file_path,
                "Line":                   f.get("line", ""),
                "Cryptographic Function": f.get("finding_type", "Config Weak Crypto"),
                "Algorithm Used":         f.get("matched_text", ""),
                "Library / Module":       "Configuration File",
                "Key Length":             "",
                "Purpose / Usage":        "Server / Application Configuration",
                "PQC Risk":               _pqc_risk_from_sev(f.get("severity", "")),
                "Crypto-Agility Support": "No",
                "Finding / Vulnerability": f.get("detail", ""),
                "_path": file_path, "_line": str(f.get("line", "")), "_mitigation": f.get("recommendation", ""),
            })
            idx += 1
    return rows


def renumber_cbom_rows(rows: list) -> list:
    """Re-assign sequential CBOM # values after combining multiple sources."""
    for i, r in enumerate(rows, start=1):
        r["# (CBOM)"] = f"CBOM #{i}"
    return rows


# ---------------------------------------------------------------------------
# Converters for new scan types (stages 9-15) → CBOM-compatible rows
# ---------------------------------------------------------------------------

def starttls_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """STARTTLS service checks (SMTP/IMAP/POP3/LDAP/FTP) → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("starttls_checks", []):
        if chk.get("error"):
            continue
        host    = chk.get("hostname", "")
        port    = chk.get("port", "")
        service = chk.get("service", "")
        system  = f"{host}:{port} ({service})"
        cipher_d = chk.get("cipher_suite") or {}
        tls_ver  = chk.get("tls_version", "")
        weak     = chk.get("weak_findings", [])
        if not weak and chk.get("starttls_supported", True):
            continue
        recs     = "\n".join(chk.get("recommendations", []))
        algo     = _extract_algo_from_cipher(cipher_d.get("name", ""))
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": f"STARTTLS / {service} Encryption",
            "Algorithm Used":         algo,
            "Library / Module":       f"{service} / TLS {tls_ver}".strip(" /"),
            "Key Length":             str(cipher_d.get("key_bits", "")),
            "Purpose / Usage":        f"{service} transport encryption",
            "PQC Risk":               "High" if weak else "Medium",
            "Crypto-Agility Support": "Partial",
            "Finding / Vulnerability": weak_str or "STARTTLS not supported",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


def ike_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """IKE/IPsec VPN scan results → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("ike_checks", []):
        if chk.get("error"):
            continue
        host    = chk.get("hostname", "")
        port    = chk.get("port", "")
        system  = f"{host}:{port}"
        version = chk.get("ike_version", "IKE")
        dh      = chk.get("preferred_dh_group") or {}
        dh_name = dh.get("name", "Unknown DH Group")
        weak    = chk.get("weak_findings", [])
        if not weak and not chk.get("ikev1_enabled"):
            continue
        recs    = "\n".join(chk.get("recommendations", []))
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        enc_algos = ", ".join(chk.get("encryption_algos", []))
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": f"{version} Key Exchange",
            "Algorithm Used":         dh_name,
            "Library / Module":       version,
            "Key Length":             enc_algos,
            "Purpose / Usage":        "VPN / IPsec key exchange",
            "PQC Risk":               "Low" if chk.get("pqc_ready") else "High",
            "Crypto-Agility Support": "No",
            "Finding / Vulnerability": weak_str or "No weak findings",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
        if chk.get("ikev1_enabled"):
            rows.append({
                "# (CBOM)":               f"CBOM #{idx}",
                "System / Application":   system,
                "File Path":              "",
                "Line":                   "",
                "Cryptographic Function": "IKEv1 Key Exchange",
                "Algorithm Used":         "IKEv1 (deprecated)",
                "Library / Module":       "IKEv1",
                "Key Length":             "",
                "Purpose / Usage":        "VPN / IPsec key exchange",
                "PQC Risk":               "High",
                "Crypto-Agility Support": "No",
                "Finding / Vulnerability": "IKEv1 is deprecated (RFC 9395); vulnerable to multiple attacks",
                "_path": system, "_line": "", "_mitigation": "Migrate to IKEv2 with PQC hybrid DH groups.",
            })
            idx += 1
    return rows


def rdp_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """RDP crypto analysis → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("rdp_checks", []):
        if chk.get("error"):
            continue
        host    = chk.get("hostname", "")
        port    = chk.get("port", "")
        system  = f"{host}:{port}"
        proto   = chk.get("security_protocol", "Unknown")
        cipher_d = chk.get("cipher_suite") or {}
        tls_ver  = chk.get("tls_version", "")
        weak    = chk.get("weak_findings", [])
        if not weak and not chk.get("classic_rdp_security"):
            continue
        recs    = "\n".join(chk.get("recommendations", []))
        algo    = _extract_algo_from_cipher(cipher_d.get("name", "")) or ("RC4" if chk.get("classic_rdp_security") else "")
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": "RDP Session Encryption",
            "Algorithm Used":         algo,
            "Library / Module":       f"RDP / {proto}",
            "Key Length":             str(cipher_d.get("key_bits", "")),
            "Purpose / Usage":        "Remote Desktop session encryption",
            "PQC Risk":               "High" if (weak or chk.get("classic_rdp_security")) else "Medium",
            "Crypto-Agility Support": "No",
            "Finding / Vulnerability": weak_str or "No weak findings",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


def smb_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """SMB version and encryption analysis → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("smb_checks", []):
        if chk.get("error"):
            continue
        host    = chk.get("hostname", "")
        port    = chk.get("port", "")
        system  = f"{host}:{port}"
        dialect = chk.get("dialect", "SMB")
        ciphers = ", ".join(chk.get("cipher_ids", []))
        weak    = chk.get("weak_findings", [])
        if not weak:
            continue
        recs    = "\n".join(chk.get("recommendations", []))
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": "SMB Session Encryption / Authentication",
            "Algorithm Used":         ciphers or ("None (SMBv1)" if chk.get("smb1_enabled") else "NTLM/Kerberos"),
            "Library / Module":       dialect,
            "Key Length":             "",
            "Purpose / Usage":        "File sharing / SMB session encryption",
            "PQC Risk":               "High" if (chk.get("smb1_enabled") or not chk.get("encryption_supported")) else "Medium",
            "Crypto-Agility Support": "No",
            "Finding / Vulnerability": weak_str or "No weak findings",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


def db_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """Database TLS analysis → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("db_checks", []):
        if chk.get("error"):
            continue
        host     = chk.get("hostname", "")
        port     = chk.get("port", "")
        service  = chk.get("service", "Database")
        system   = f"{host}:{port} ({service})"
        cipher_d = chk.get("cipher_suite") or {}
        tls_ver  = chk.get("tls_version", "")
        weak     = chk.get("weak_findings", [])
        if not weak and chk.get("tls_supported", True):
            continue
        recs     = "\n".join(chk.get("recommendations", []))
        algo     = _extract_algo_from_cipher(cipher_d.get("name", ""))
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": f"{service} Transport Encryption",
            "Algorithm Used":         algo or ("None" if not chk.get("tls_supported") else ""),
            "Library / Module":       f"{service} / TLS {tls_ver}".strip(" /"),
            "Key Length":             str(cipher_d.get("key_bits", "")),
            "Purpose / Usage":        f"{service} data-in-transit encryption",
            "PQC Risk":               "High" if (not chk.get("tls_supported") or weak) else "Medium",
            "Crypto-Agility Support": "No",
            "Finding / Vulnerability": weak_str or ("No TLS" if not chk.get("tls_supported") else "No weak findings"),
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


def snmp_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """SNMP version/auth analysis → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("snmp_checks", []):
        if chk.get("error"):
            continue
        host    = chk.get("hostname", "")
        port    = chk.get("port", "")
        system  = f"{host}:{port}"
        weak    = chk.get("weak_findings", [])
        if not weak:
            continue
        recs    = "\n".join(chk.get("recommendations", []))
        versions = []
        if chk.get("v1_enabled"):  versions.append("SNMPv1")
        if chk.get("v2c_enabled"): versions.append("SNMPv2c")
        if chk.get("v3_enabled"):  versions.append("SNMPv3")
        ver_str = ", ".join(versions) or "Unknown"
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": "SNMP Authentication / Privacy",
            "Algorithm Used":         "Community String (MD5/SHA)" if (chk.get("v1_enabled") or chk.get("v2c_enabled")) else "USM",
            "Library / Module":       f"SNMP ({ver_str})",
            "Key Length":             "",
            "Purpose / Usage":        "Network device management authentication",
            "PQC Risk":               "High" if (chk.get("v1_enabled") or chk.get("v2c_enabled")) else "Medium",
            "Crypto-Agility Support": "No",
            "Finding / Vulnerability": weak_str or "No weak findings",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


def kerberos_to_cbom_rows(data: dict, start_idx: int = 1) -> list:
    """Kerberos KDC encryption type analysis → CBOM rows."""
    rows = []
    idx = start_idx
    for chk in data.get("kerberos_checks", []):
        if chk.get("error"):
            continue
        host    = chk.get("hostname", "")
        port    = chk.get("port", "")
        system  = f"{host}:{port}"
        etypes  = ", ".join(chk.get("etype_names", []))
        weak    = chk.get("weak_findings", [])
        if not weak:
            continue
        recs    = "\n".join(chk.get("recommendations", []))
        weak_str = "; ".join(f"{w.get('type')} [{w.get('severity')}]: {w.get('detail','')}" for w in weak)
        rows.append({
            "# (CBOM)":               f"CBOM #{idx}",
            "System / Application":   system,
            "File Path":              "",
            "Line":                   "",
            "Cryptographic Function": "Kerberos Ticket Encryption",
            "Algorithm Used":         etypes or "Unknown",
            "Library / Module":       "Kerberos KDC",
            "Key Length":             "",
            "Purpose / Usage":        "Kerberos authentication / ticket encryption",
            "PQC Risk":               "High" if (chk.get("rc4_enabled") or chk.get("des_enabled")) else "Medium",
            "Crypto-Agility Support": "No",
            "Finding / Vulnerability": weak_str or "No weak findings",
            "_path": system, "_line": "", "_mitigation": recs,
        })
        idx += 1
    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(
    sbom_path, grype_path, cbom_path, mini_path, out_path,
    tls_path=None, ssh_path=None, cert_path=None, config_path=None,
    starttls_path=None, ike_path=None, rdp_path=None, smb_path=None,
    db_path=None, snmp_path=None, kerberos_path=None,
):
    # ---- load scan data ----
    sbom  = load_json(sbom_path)
    grype = load_json(grype_path)
    cbom  = load_json(cbom_path)
    _mini = load_json(mini_path)

    grype_worst = parse_grype_worst_severity_by_pkg(grype)
    sbom_rows   = parse_sbom_rows(sbom, grype_worst)
    cbom_rows   = parse_cbom_rows(cbom)

    # ---- load all optional scan sources ----
    tls_data      = load_json(tls_path)      if tls_path      else None
    ssh_data      = load_json(ssh_path)      if ssh_path      else None
    cert_data     = load_json(cert_path)     if cert_path     else None
    config_data   = load_json(config_path)   if config_path   else None
    starttls_data = load_json(starttls_path) if starttls_path else None
    ike_data      = load_json(ike_path)      if ike_path      else None
    rdp_data      = load_json(rdp_path)      if rdp_path      else None
    smb_data      = load_json(smb_path)      if smb_path      else None
    db_data       = load_json(db_path)       if db_path       else None
    snmp_data     = load_json(snmp_path)     if snmp_path     else None
    kerberos_data = load_json(kerberos_path) if kerberos_path else None

    # ---- combine all scan sources into one CBOM row list ----
    extra_cbom: list = []
    def _add(fn, src):
        if src:
            extra_cbom.extend(fn(src, start_idx=len(cbom_rows) + len(extra_cbom) + 1))

    _add(tls_to_cbom_rows,      tls_data)
    _add(ssh_to_cbom_rows,      ssh_data)
    _add(cert_to_cbom_rows,     cert_data)
    _add(config_to_cbom_rows,   config_data)
    _add(starttls_to_cbom_rows, starttls_data)
    _add(ike_to_cbom_rows,      ike_data)
    _add(rdp_to_cbom_rows,      rdp_data)
    _add(smb_to_cbom_rows,      smb_data)
    _add(db_to_cbom_rows,       db_data)
    _add(snmp_to_cbom_rows,     snmp_data)
    _add(kerberos_to_cbom_rows, kerberos_data)

    all_cbom_rows = renumber_cbom_rows(cbom_rows + extra_cbom)

    # ---- risk rows from full combined CBOM ----
    rr_rows, ra_rows = risk_rows_from_cbom(all_cbom_rows)

    inventory_rows = parse_inventory_rows(
        sbom, grype_worst, all_cbom_rows,
        mini_pqc=_mini,
        tls_data=tls_data,       ssh_data=ssh_data,
        cert_data=cert_data,     config_data=config_data,
        starttls_data=starttls_data, ike_data=ike_data,
        rdp_data=rdp_data,       smb_data=smb_data,
        db_data=db_data,         snmp_data=snmp_data,
        kerberos_data=kerberos_data,
    )

    # ---- start from template (preserves static sheets + all styling) ----
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Reference template not found: {TEMPLATE_PATH}\n"
            "Ensure 'bin/pqc_migration/BUKUKERJA_BENGKEL MIGRASI PQC 2025.xlsx' is present."
        )
    shutil.copy2(TEMPLATE_PATH, out_path)
    wb = load_workbook(out_path)

    # ---- 0_Inventory (auto-populated from all scan sources incl. TLS/SSH/Cert/Config) ----
    populate_sheet(wb["0_Inventory"], INVENTORY_HEADERS, inventory_rows)

    # ---- 1_SBOM ----
    populate_sheet(wb["1_SBOM"], SBOM_HEADERS, sbom_rows)

    # ---- 2_CBOM (combined: Semgrep + TLS + SSH + Cert + Config findings) ----
    populate_sheet(wb["2_CBOM"], CBOM_HEADERS, all_cbom_rows)

    # ---- 3_RiskRegister (risk entries from all combined CBOM sources) ----
    populate_sheet(wb["3_RiskRegister"], RR_HEADERS, rr_rows)

    # ---- 4_RiskAssessment (assessments from all combined CBOM sources) ----
    populate_sheet(wb["4_RiskAssessment"], RA_HEADERS, ra_rows)

    # ---- 5_RiskMatrix, 6_ProtocolCryptoMap, 00_ReadMe ----
    # Kept as-is from the template (static reference content).

    wb.save(out_path)


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--sbom",   required=True)
    ap.add_argument("--grype",  required=True)
    ap.add_argument("--cbom",   required=True)
    ap.add_argument("--mini",   required=True)
    ap.add_argument("--out",    required=True)
    ap.add_argument("--tls",    default=None, help="TLS scan JSON (Stage 5)")
    ap.add_argument("--ssh",    default=None, help="SSH scan JSON (Stage 6)")
    ap.add_argument("--certs",  default=None, help="Cert file scan JSON (Stage 7)")
    ap.add_argument("--config", default=None, help="Config file scan JSON (Stage 8)")
    args = ap.parse_args()
    main(args.sbom, args.grype, args.cbom, args.mini, args.out,
         tls_path=args.tls, ssh_path=args.ssh,
         cert_path=args.certs, config_path=args.config)
