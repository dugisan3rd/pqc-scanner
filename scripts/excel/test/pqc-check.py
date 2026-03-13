#!/usr/bin/env python3
"""
Full-ish PQC Readiness Audit (Admin/Sudo) + XLSX Output

Outputs:
- pqc_readiness_report.json  (raw evidence)
- pqc_readiness_report.txt   (human summary)
- pqc_readiness_report.xlsx  (structured workbook)

Run (Linux/macOS):
  sudo python3 pqc_full_audit_xlsx.py --install-tools

Run (Windows Admin PowerShell):
  python .\pqc_full_audit_xlsx.py --install-tools

Dependencies:
  pip install psutil openpyxl
"""

from __future__ import annotations

import argparse
import json
import os
import platform
import re
import shutil
import socket
import subprocess
import sys
import time
from typing import Any, Dict, List, Optional, Tuple

try:
    import psutil  # type: ignore
except ImportError:
    print("Missing dependency: psutil. Install with: pip install psutil", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Missing dependency: openpyxl. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# --------------------- helpers ---------------------

def now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%S%z")

def which(cmd: str) -> Optional[str]:
    return shutil.which(cmd)

def run_cmd(cmd: List[str], timeout: int = 120) -> Tuple[int, str, str]:
    try:
        p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return p.returncode, (p.stdout or "").strip(), (p.stderr or "").strip()
    except FileNotFoundError:
        return 127, "", f"not found: {cmd[0]}"
    except subprocess.TimeoutExpired:
        return 124, "", "timeout"
    except Exception as e:
        return 1, "", f"error: {e}"

def is_admin() -> bool:
    sysname = platform.system().lower()
    if "windows" in sysname:
        try:
            import ctypes  # type: ignore
            return bool(ctypes.windll.shell32.IsUserAnAdmin())
        except Exception:
            return False
    else:
        try:
            return os.geteuid() == 0  # type: ignore[attr-defined]
        except Exception:
            return False

def detect_pkg_manager() -> Optional[str]:
    sysname = platform.system().lower()
    if "windows" in sysname:
        if which("winget"):
            return "winget"
        if which("choco"):
            return "choco"
        return None
    if "darwin" in sysname:
        return "brew" if which("brew") else None
    for pm in ["apt-get", "dnf", "yum", "pacman", "zypper", "apk"]:
        if which(pm):
            return pm
    return None

def install_tools(tools: List[str]) -> Dict[str, Any]:
    pm = detect_pkg_manager()
    out: Dict[str, Any] = {"pkg_manager": pm, "requested": tools, "actions": [], "notes": []}

    if not pm:
        out["notes"].append("No supported package manager detected. Install tools manually.")
        return out

    def already(name: str) -> bool:
        return which(name) is not None

    missing = [t for t in tools if not already(t)]
    out["missing"] = missing
    if not missing:
        out["notes"].append("All requested tools already present.")
        return out

    if not is_admin():
        out["notes"].append("Not running as admin/root. Tool installation may fail.")

    # package-manager specific
    if pm == "apt-get":
        run_cmd(["apt-get", "update"], timeout=600)
        cmd = ["apt-get", "install", "-y"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    if pm in ("dnf", "yum"):
        cmd = [pm, "install", "-y"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    if pm == "pacman":
        cmd = ["pacman", "-Sy", "--noconfirm"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    if pm == "zypper":
        cmd = ["zypper", "--non-interactive", "install"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    if pm == "apk":
        cmd = ["apk", "add"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    if pm == "brew":
        cmd = ["brew", "install"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    if pm == "winget":
        actions = []
        for t in missing:
            actions.append(["winget", "install", "--silent", "--accept-package-agreements",
                            "--accept-source-agreements", t])
        out["actions"] = [{"cmd": a} for a in actions]
        for a in actions:
            run_cmd(a, timeout=900)
        return out

    if pm == "choco":
        cmd = ["choco", "install", "-y"] + missing
        rc, stdout, stderr = run_cmd(cmd, timeout=900)
        out["actions"].append({"cmd": cmd, "rc": rc})
        out["stdout"] = stdout[-6000:]
        out["stderr"] = stderr[-6000:]
        return out

    out["notes"].append(f"Unsupported package manager handler: {pm}")
    return out

def get_os_info() -> Dict[str, str]:
    return {
        "system": platform.system(),
        "release": platform.release(),
        "version": platform.version(),
        "machine": platform.machine(),
        "python": platform.python_version(),
        "node": platform.node() or socket.gethostname(),
    }

def read_file(path: str, max_bytes: int = 250_000) -> Optional[str]:
    try:
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read(max_bytes)
    except Exception:
        return None

def list_listening_ports() -> List[Dict[str, Any]]:
    res = []
    for c in psutil.net_connections(kind="inet"):
        if c.status != psutil.CONN_LISTEN:
            continue
        laddr = c.laddr
        if not laddr:
            continue
        proto = "tcp" if c.type == socket.SOCK_STREAM else "udp"
        ip = getattr(laddr, "ip", laddr[0] if isinstance(laddr, tuple) else str(laddr))
        port = getattr(laddr, "port", laddr[1] if isinstance(laddr, tuple) else 0)
        pid = c.pid
        pname = None
        pexe = None
        if pid:
            try:
                p = psutil.Process(pid)
                pname = p.name()
                pexe = p.exe() if hasattr(p, "exe") else None
            except Exception:
                pass
        res.append({"proto": proto, "ip": ip, "port": port, "pid": pid, "process": pname, "exe": pexe})
    res.sort(key=lambda x: (x["proto"], x["port"], x["ip"]))
    return res

def get_versions() -> Dict[str, Any]:
    tools = {}
    for tool, cmd in {
        "nmap": ["nmap", "--version"],
        "openssl": ["openssl", "version", "-a"],
        "ssh": ["ssh", "-V"],
        "sshd": ["sshd", "-T"],
        "java": ["java", "-version"],
    }.items():
        path = which(tool)
        if not path:
            tools[tool] = {"present": False, "path": None, "first_line": None, "raw": None}
            continue
        rc, out, err = run_cmd(cmd, timeout=30)
        raw = out if out else err
        first = None
        if raw:
            for line in raw.splitlines():
                if line.strip():
                    first = line.strip()
                    break
        tools[tool] = {"present": True, "path": path, "rc": rc, "first_line": first, "raw": raw[-8000:]}
    return tools

def nmap_tls_enum(target: str, ports: List[int]) -> Dict[str, Any]:
    if not which("nmap"):
        return {"ok": False, "error": "nmap not found"}
    if not ports:
        return {"ok": True, "note": "No TCP ports selected for TLS enumeration."}
    port_str = ",".join(str(p) for p in sorted(set(ports)))
    cmd = ["nmap", "-Pn", "--script", "ssl-enum-ciphers", "-p", port_str, target]
    rc, out, err = run_cmd(cmd, timeout=900)
    return {"ok": rc == 0, "cmd": cmd, "rc": rc, "stdout": out[-250000:], "stderr": err[-25000:]}

def nmap_ssh_enum(target: str, ports: List[int]) -> Dict[str, Any]:
    if not which("nmap"):
        return {"ok": False, "error": "nmap not found"}
    ssh_ports = [p for p in ports if p in (22, 2222, 22222)]
    if not ssh_ports:
        return {"ok": True, "note": "No common SSH port detected to scan (22/2222/22222)."}
    port_str = ",".join(str(p) for p in sorted(set(ssh_ports)))
    cmd = ["nmap", "-Pn", "--script", "ssh2-enum-algos", "-p", port_str, target]
    rc, out, err = run_cmd(cmd, timeout=600)
    return {"ok": rc == 0, "cmd": cmd, "rc": rc, "stdout": out[-250000:], "stderr": err[-25000:]}

def extract_sshd_policy() -> Dict[str, Any]:
    data: Dict[str, Any] = {"sshd_T": None, "sshd_config": None, "notes": []}
    if which("sshd"):
        rc, out, err = run_cmd(["sshd", "-T"], timeout=20)
        data["sshd_T"] = {"rc": rc, "stdout": out[-80000:], "stderr": err[-8000:]}
    else:
        data["notes"].append("sshd not found in PATH.")
    cfg_paths = (
        [r"C:\ProgramData\ssh\sshd_config", r"C:\Program Files\OpenSSH\sshd_config"]
        if "windows" in platform.system().lower()
        else ["/etc/ssh/sshd_config", "/usr/local/etc/sshd_config"]
    )
    for p in cfg_paths:
        t = read_file(p)
        if t:
            data["sshd_config"] = {"path": p, "excerpt": t[:150000]}
            break
    if not data["sshd_config"]:
        data["notes"].append("sshd_config not found/readable in common locations.")
    return data

def extract_tls_termination_configs() -> Dict[str, Any]:
    res: Dict[str, Any] = {"nginx": None, "apache": None, "haproxy": None}
    if which("nginx"):
        rc, out, err = run_cmd(["nginx", "-T"], timeout=40)
        res["nginx"] = {"present": True, "rc": rc, "stdout": out[-200000:], "stderr": err[-12000:]}
    else:
        res["nginx"] = {"present": False}
    apache_bins = ["apachectl", "httpd", "apache2ctl"]
    apache_bin = next((b for b in apache_bins if which(b)), None)
    if apache_bin:
        rc, out, err = run_cmd([apache_bin, "-S"], timeout=40)
        res["apache"] = {"present": True, "bin": apache_bin, "rc": rc, "stdout": out[-120000:], "stderr": err[-12000:]}
    else:
        res["apache"] = {"present": False}
    if which("haproxy"):
        rc, out, err = run_cmd(["haproxy", "-vv"], timeout=30)
        res["haproxy"] = {"present": True, "rc": rc, "stdout": out[-120000:], "stderr": err[-12000:]}
    else:
        res["haproxy"] = {"present": False}
    return res

def pqc_signals_from_text(text: str) -> Dict[str, Any]:
    hits = []
    keywords = [
        "sntrup761x25519", "mlkem", "kyber", "ntru", "dilithium", "mldsa",
        "pqc", "hybrid", "post-quantum", "post quantum"
    ]
    lt = (text or "").lower()
    for k in keywords:
        if k in lt:
            hits.append(k)
    return {"pqc_keywords_found": sorted(set(hits))}

def score_readiness(report: Dict[str, Any]) -> Dict[str, Any]:
    score = 0
    reasons = []
    versions = report.get("versions", {})
    if versions.get("nmap", {}).get("present"):
        score += 1; reasons.append("nmap present (capability enumeration)")
    if versions.get("openssl", {}).get("present"):
        score += 1; reasons.append("openssl present (crypto stack visibility)")
    if report.get("sshd", {}).get("sshd_T", {}).get("rc") == 0:
        score += 1; reasons.append("sshd policy extractable (configurable)")
    cfg = report.get("tls_termination_configs", {})
    if cfg.get("nginx", {}).get("present") or cfg.get("haproxy", {}).get("present"):
        score += 1; reasons.append("central TLS termination likely (nginx/haproxy detected)")
    if report.get("pqc_signals", {}).get("pqc_keywords_found"):
        score += 1; reasons.append("hybrid/PQC keywords detected (heuristic)")
    score = max(0, min(score, 5))
    label = ["Very Low", "Low", "Moderate", "Good", "Strong", "Very Strong"][score]
    return {"score_0_to_5": score, "label": label, "reasons": reasons, "note": "Crypto-agility signals, not proof of PQC adoption."}

def write_txt_report(path: str, data: Dict[str, Any]) -> None:
    lines = []
    lines.append("=" * 78)
    lines.append("PQC READINESS TECHNICAL AUDIT (LOCAL, ADMIN)")
    lines.append("=" * 78)
    lines.append(f"Generated: {data.get('generated_at')}")
    osinfo = data.get("os", {})
    lines.append(f"Host: {osinfo.get('node')} | OS: {osinfo.get('system')} {osinfo.get('release')} ({osinfo.get('machine')})")
    lines.append(f"Admin/root: {data.get('admin')}")
    lines.append("")
    lines.append("Listening services:")
    for s in data.get("listening_services", []):
        lines.append(f"  - {s['proto']} {s['ip']}:{s['port']} pid={s.get('pid')} proc={s.get('process')}")
    lines.append("")
    lines.append("PQC signals (heuristic): " + str(data.get("pqc_signals", {}).get("pqc_keywords_found")))
    sc = data.get("readiness_score", {})
    lines.append(f"Crypto-agility score: {sc.get('score_0_to_5')}/5 ({sc.get('label')})")
    for r in sc.get("reasons", []) or []:
        lines.append(f"  * {r}")
    lines.append("")
    lines.append("Notes:")
    lines.append("  - RSA/ECDSA/ECDH are classical (future Shor risk).")
    lines.append("  - PQC readiness practically requires: inventory + data lifetime + client constraints + migration plan.")
    lines.append("=" * 78)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# --------------------- XLSX helpers ---------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

def autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            s = str(cell.value)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)

def add_table(ws, start_row: int, start_col: int, end_row: int, end_col: int, name: str) -> None:
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def parse_nmap_ssl_enum(text: str) -> List[Dict[str, Any]]:
    """
    Very light parser for nmap ssl-enum-ciphers output:
    Captures: port, tls_version blocks, cipher lines with strength grades.
    """
    rows: List[Dict[str, Any]] = []
    if not text:
        return rows
    current_port = None
    current_tls = None

    # Port header example: "443/tcp open  https"
    port_re = re.compile(r"^(\d+)/tcp\s+open", re.IGNORECASE)
    # TLS block example: "|   TLSv1.2:"
    tls_re = re.compile(r"^\|\s+(TLSv1\.\d|TLSv1\.\d+|TLSv1\.3|SSLv3)\s*:", re.IGNORECASE)
    # Cipher line example: "|     ciphers:"
    cipher_re = re.compile(r"^\|\s+([A-Z0-9_\-]+)\s+\(([^)]+)\)\s+\-\s+([A-F])", re.IGNORECASE)

    for line in text.splitlines():
        m = port_re.search(line)
        if m:
            current_port = int(m.group(1))
            current_tls = None
            continue
        m = tls_re.search(line)
        if m:
            current_tls = m.group(1).upper()
            continue
        m = cipher_re.search(line)
        if m and current_port and current_tls:
            cipher = m.group(1)
            meta = m.group(2)
            grade = m.group(3).upper()
            rows.append({
                "port": current_port,
                "tls_version": current_tls,
                "cipher": cipher,
                "details": meta,
                "grade": grade
            })
    return rows

def parse_nmap_ssh_enum(text: str) -> List[Dict[str, Any]]:
    """
    Light parser for nmap ssh2-enum-algos output:
    Extracts KEX, hostkey, ciphers, MACs.
    """
    rows: List[Dict[str, Any]] = []
    if not text:
        return rows
    current_port = None
    port_re = re.compile(r"^(\d+)/tcp\s+open\s+ssh", re.IGNORECASE)
    section = None
    sec_re = re.compile(r"^\|\s+(kex_algorithms|server_host_key_algorithms|encryption_algorithms|mac_algorithms|compression_algorithms):", re.IGNORECASE)
    algo_re = re.compile(r"^\|\s+([^\s].+)$")

    for line in text.splitlines():
        m = port_re.search(line)
        if m:
            current_port = int(m.group(1))
            section = None
            continue
        m = sec_re.search(line)
        if m:
            section = m.group(1).lower()
            continue
        if section and current_port and line.strip().startswith("|"):
            m = algo_re.search(line)
            if m:
                val = m.group(1).strip()
                if val.endswith(":"):
                    continue
                # ignore bullets like "|" only
                if val in ("", "ciphers:", "macs:"):
                    continue
                rows.append({"port": current_port, "category": section, "algorithm": val})
    return rows

def write_xlsx(path: str, report: Dict[str, Any]) -> None:
    wb = Workbook()

    def add_table_if_data(ws, name: str, start_row: int, start_col: int, end_col: int):
        # Only add a table if there is at least 1 data row (header + 1)
        if ws.max_row >= start_row + 1:
            add_table(ws, start_row, start_col, ws.max_row, end_col, name)

    # --- Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    ws["A1"].fill = HEADER_FILL; ws["B1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT; ws["B1"].font = HEADER_FONT
    ws.freeze_panes = "A2"

    osinfo = report.get("os", {})
    sc = report.get("readiness_score", {})
    summary_rows = [
        ("Generated At", report.get("generated_at")),
        ("Host", osinfo.get("node")),
        ("OS", f"{osinfo.get('system')} {osinfo.get('release')}"),
        ("Machine", osinfo.get("machine")),
        ("Admin/Root", str(report.get("admin"))),
        ("Scan Target", report.get("scanned_target")),
        ("Ports Scanned", ", ".join(map(str, report.get("scanned_ports", [])))),
        ("PQC Signals (heuristic)", ", ".join(report.get("pqc_signals", {}).get("pqc_keywords_found", []))),
        ("Crypto-Agility Score", f"{sc.get('score_0_to_5')}/5"),
        ("Score Label", sc.get("label")),
    ]
    for k, v in summary_rows:
        ws.append([k, v])

    autosize(ws, max_width=80)
    add_table_if_data(ws, "SummaryTable", 1, 1, 2)

    # --- Tools
    ws = wb.create_sheet("Tools")
    headers = ["Tool", "Present", "Path", "First Line", "RC"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT

    versions = report.get("versions", {}) or {}
    for tool, v in versions.items():
        ws.append([tool, v.get("present"), v.get("path"), v.get("first_line"), v.get("rc")])

    ws.freeze_panes = "A2"
    autosize(ws, max_width=90)
    add_table_if_data(ws, "ToolsTable", 1, 1, len(headers))

    # --- Listening Services
    ws = wb.create_sheet("ListeningServices")
    headers = ["Proto", "IP", "Port", "PID", "Process", "EXE"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT

    for s in report.get("listening_services", []) or []:
        ws.append([s.get("proto"), s.get("ip"), s.get("port"), s.get("pid"), s.get("process"), s.get("exe")])

    ws.freeze_panes = "A2"
    autosize(ws, max_width=70)
    add_table_if_data(ws, "ListenTable", 1, 1, len(headers))

    # --- TLS Enum (Parsed)
    ws = wb.create_sheet("TLS_Enum")
    headers = ["Port", "TLS Version", "Cipher", "Details", "Grade"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT

    tls_text = (report.get("nmap_tls_enum", {}) or {}).get("stdout", "") or ""
    tls_rows = parse_nmap_ssl_enum(tls_text)
    for r in tls_rows:
        ws.append([r["port"], r["tls_version"], r["cipher"], r["details"], r["grade"]])

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = WRAP
    autosize(ws, max_width=80)
    add_table_if_data(ws, "TLSEnumTable", 1, 1, 5)

    # --- SSH Enum (Parsed)
    ws = wb.create_sheet("SSH_Enum")
    headers = ["Port", "Category", "Algorithm"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT

    ssh_text = (report.get("nmap_ssh_enum", {}) or {}).get("stdout", "") or ""
    ssh_rows = parse_nmap_ssh_enum(ssh_text)
    for r in ssh_rows:
        ws.append([r["port"], r["category"], r["algorithm"]])

    ws.freeze_panes = "A2"
    autosize(ws, max_width=100)
    add_table_if_data(ws, "SSHEngTable", 1, 1, 3)

    # --- Evidence (Raw excerpts)
    ws = wb.create_sheet("Evidence_Raw")
    headers = ["Item", "Excerpt"]
    ws.append(headers)
    ws["A1"].fill = HEADER_FILL; ws["B1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT; ws["B1"].font = HEADER_FONT
    ws.freeze_panes = "A2"

    def add_blob(name: str, blob: str, limit: int = 60000):
        if blob:
            ws.append([name, blob[:limit]])

    add_blob("nmap_tls_enum.stdout", tls_text)
    add_blob("nmap_tls_enum.stderr", (report.get("nmap_tls_enum", {}) or {}).get("stderr", ""))
    add_blob("nmap_ssh_enum.stdout", ssh_text)
    add_blob("nmap_ssh_enum.stderr", (report.get("nmap_ssh_enum", {}) or {}).get("stderr", ""))

    sshd = report.get("sshd", {}) or {}
    add_blob("sshd -T stdout", (sshd.get("sshd_T", {}) or {}).get("stdout", ""))
    add_blob("sshd_config excerpt", (sshd.get("sshd_config", {}) or {}).get("excerpt", ""))

    cfg = report.get("tls_termination_configs", {}) or {}
    add_blob("nginx -T stdout", (cfg.get("nginx", {}) or {}).get("stdout", ""))
    add_blob("apache -S stdout", (cfg.get("apache", {}) or {}).get("stdout", ""))
    add_blob("haproxy -vv stdout", (cfg.get("haproxy", {}) or {}).get("stdout", ""))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        row[1].alignment = WRAP

    add_table_if_data(ws, "EvidenceTable", 1, 1, 2)

    wb.save(path)


# --------------------- main ---------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", default="127.0.0.1", help="Target to scan (default localhost)")
    ap.add_argument("--install-tools", action="store_true", help="Attempt to install nmap/openssl/curl/git (admin/root recommended)")
    ap.add_argument("--out-json", default="pqc_readiness_report.json")
    ap.add_argument("--out-txt", default="pqc_readiness_report.txt")
    ap.add_argument("--out-xlsx", default="pqc_readiness_report.xlsx")
    ap.add_argument("--max-ports", type=int, default=50, help="Max ports to scan via nmap scripts")
    args = ap.parse_args()

    inst = install_tools(["nmap", "openssl", "curl", "git"]) if args.install_tools else {"note": "Tool install skipped."}

    osinfo = get_os_info()
    listening = list_listening_ports()
    tcp_ports = sorted({s["port"] for s in listening if s["proto"] == "tcp"})

    priority = [443, 8443, 9443, 22, 2222, 636, 993, 995, 465, 587, 3306, 5432, 9200, 5601]
    ordered: List[int] = []
    for p in priority:
        if p in tcp_ports and p not in ordered:
            ordered.append(p)
    for p in tcp_ports:
        if p not in ordered:
            ordered.append(p)
    ordered = ordered[: max(1, args.max_ports)]

    versions = get_versions()
    tls_enum = nmap_tls_enum(args.target, [p for p in ordered if p != 22])
    ssh_enum = nmap_ssh_enum(args.target, ordered)
    sshd = extract_sshd_policy()
    tls_cfg = extract_tls_termination_configs()

    combined_text = ""
    combined_text += (tls_enum.get("stdout") or "") + "\n"
    combined_text += (ssh_enum.get("stdout") or "") + "\n"
    combined_text += (sshd.get("sshd_T", {}) or {}).get("stdout", "") + "\n"
    combined_text += (tls_cfg.get("nginx", {}) or {}).get("stdout", "") + "\n"
    pqc_signals = pqc_signals_from_text(combined_text)

    report: Dict[str, Any] = {
        "generated_at": now_iso(),
        "admin": is_admin(),
        "os": osinfo,
        "install": inst,
        "versions": versions,
        "listening_services": listening,
        "scanned_target": args.target,
        "scanned_ports": ordered,
        "nmap_tls_enum": tls_enum,
        "nmap_ssh_enum": ssh_enum,
        "sshd": sshd,
        "tls_termination_configs": tls_cfg,
        "pqc_signals": pqc_signals,
    }
    report["readiness_score"] = score_readiness(report)

    with open(args.out_json, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    write_txt_report(args.out_txt, report)
    write_xlsx(args.out_xlsx, report)

    print(f"[OK] JSON : {args.out_json}")
    print(f"[OK] TXT  : {args.out_txt}")
    print(f"[OK] XLSX : {args.out_xlsx}")
    print(f"[INFO] Admin/root: {report['admin']}")
    print(f"[INFO] PQC signals (heuristic): {report['pqc_signals'].get('pqc_keywords_found')}")
    print(f"[INFO] Crypto-agility score: {report['readiness_score'].get('score_0_to_5')}/5 ({report['readiness_score'].get('label')})")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())