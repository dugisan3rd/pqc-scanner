#!/usr/bin/env python3
"""
CBOM Report Generator
=====================
Generates a multi-sheet Excel workbook from cryptographic and software bill-of-materials data.

Input files:
  - cbom.json    Semgrep JSON with metadata.cbom fields
  - sbom.json    Syft SBOM JSON
  - grype.json   Grype vulnerability scan JSON

Output:
  - cbom_output.xlsx  (5 sheets: CBOM, CBOM_Summary, SBOM_Grype, PQC_Risk, PQC_RiskMatrix)
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

INPUT_CBOM    = Path("cbom.json")
INPUT_SBOM    = Path("sbom.json")
INPUT_GRYPE   = Path("grype.json")
INPUT_MINIPQC = Path("minipqc.json")
OUTPUT_XLSX   = Path("cbom_output.xlsx")

# ---------------------------------------------------------------------------
# Column schemas
# ---------------------------------------------------------------------------

CBOM_COLS = [
    "# (CBOM)",
    "System / Application",
    "Cryptographic Function",
    "Algorithm Used",
    "Library / Module",
    "File / Location",
    "Key Length",
    "Purpose / Usage",
    "Crypto-Agility Support",
]

SBOM_COLS = [
    "#",
    "System / Application",
    "Purpose / Usage",
    "URL",
    "Services Mode",
    "Target Customer",
    "Software Component",
    "Vulnerability Status",
    "Third-party Modules",
    "External APIs or Services",
    "Critical Level",
    "Data Category",
    "Is the application/system currently in use?",
    "Application/System Developer",
    "Vendor's Name",
    "Does the agency have expertise?",
    "Does the agency have a special budget allocation?",
    "Link to CBOM",
]

PQC_RISK_COLS = [
    "#",
    "System / Application",
    "Asset Type",
    "Algorithm Used",
    "Purpose / Usage",
    "Severity",
    "Risk",
    "Risk Owner",
]

PQC_MATRIX_COLS = [
    "#",
    "System/Hardware/Software Name",
    "Cryptographic Algorithm",
    "Risk",
    "Root Cause",
    "Impact",
    "Likelihood",
    "Risk Score",
    "Risk Level",
    "Existing Controls",
    "Mitigation Plan",
]

# ---------------------------------------------------------------------------
# Algorithm classification sets
# ---------------------------------------------------------------------------
# These sets use EXACT token matching (see `token_matches`) to avoid false
# positives from substring collision (e.g. "SECP" matching "ACCEPT").

# Asymmetric algorithms fully broken by Shor's algorithm on a CRQC.
# Source: NIST SP 800-131A Rev 2, NIST IR 8105
SHOR_VULNERABLE_ALGORITHMS = frozenset({
    "RSA", "DSA", "DH", "ECDH", "ECDSA", "ED25519", "ED448",
    "ECDHE", "ECDHE-RSA", "ECDHE-ECDSA",
    # Named curves (exact tokens as they appear in CBOM algorithm fields)
    "P-192", "P-256", "P-384", "P-521",
    "SECP256K1", "SECP384R1", "SECP521R1",
    "X25519", "X448",
})

# Cryptographic functions that imply asymmetric usage regardless of named algo.
# "kernel primitive" removed — not a recognised cryptographic category.
SHOR_VULNERABLE_FUNCTIONS = frozenset({
    "digital signature",
    "key exchange",
    "key encapsulation",
    "asymmetric encryption",
    "public key encryption",
})

# --------------------------------------------------------------------------
# Quantum-weakened symmetric / hash algorithms.
# Grover's algorithm halves effective bit-strength, but NIST does NOT require
# migration away from these — doubling key/digest size restores security.
# Source: NIST SP 800-175B Rev 1, NIST FAQ on PQC
#
# Sub-classification:
#   CLASSICALLY_BROKEN  — already insecure without quantum computers.
#                         Flag primarily as "legacy/broken", quantum is moot.
#   GROVER_ADEQUATE     — already meet post-quantum security at current sizes
#                         (AES-256, SHA-384, SHA-512, BLAKE2b-512, ChaCha20).
#   GROVER_NEEDS_UPGRADE— quantum-weakened AND current standard sizes may fall
#                         short (AES-128 → needs AES-256; SHA-256 → prefer SHA-384+).

CLASSICALLY_BROKEN_ALGORITHMS = frozenset({
    "MD5", "SHA1", "SHA-1",
    "DES", "3DES", "TDEA",
    "RC4", "RC2",
    "BLOWFISH",       # no longer recommended; max 64-bit block
})

GROVER_ADEQUATE_ALGORITHMS = frozenset({
    # Already ≥256-bit security post-Grover — no action needed for PQC
    "AES-256", "AES256",
    "SHA-384", "SHA384",
    "SHA-512", "SHA512", "SHA-512/256", "SHA3-384", "SHA3-512",
    "BLAKE2B-512", "BLAKE3",
    "CHACHA20-POLY1305", "XCHACHA20",
})

GROVER_NEEDS_UPGRADE_ALGORITHMS = frozenset({
    # Grover halves effective strength; upgrade key/digest size recommended
    "AES-128", "AES128", "AES-192", "AES192",
    "AES",              # unspecified key length — flag for review
    "SHA-256", "SHA256", "SHA-224", "SHA224",
    "SHA3-256", "SHA3-224",
    "BLAKE2S", "BLAKE2B-256",
    "POLY1305",         # MAC only, but flag for completeness
    "CHACHA20",         # without POLY1305 tag — key length context needed
    "MCRYPT",           # generic; key length unknown
})

HASH_FUNCTIONS    = frozenset({"hashing", "message digest", "mac", "hmac"})
SYMMETRIC_FUNCTIONS = frozenset({"encryption/decryption", "symmetric encryption", "stream cipher", "block cipher"})

# ---------------------------------------------------------------------------
# General helpers
# ---------------------------------------------------------------------------

def coalesce(*values: Any, default: str = "Unknown") -> str:
    """Return the first non-None, non-blank string value, or *default*."""
    for v in values:
        if v is not None and str(v).strip():
            return str(v)
    return default


def is_ipv4(value: str) -> bool:
    """Return True if *value* looks like an IPv4 address."""
    return bool(re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", (value or "").strip()))


def severity_rank(severity: str) -> int:
    return {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}.get(
        (severity or "").upper(), 0
    )


def max_severity(severities: List[str]) -> str:
    best, best_rank = "", 0
    for s in severities:
        r = severity_rank(s)
        if r > best_rank:
            best_rank, best = r, s
    return best.upper() if best_rank > 0 else "None"


def deduplicated(items: List[str]) -> List[str]:
    seen: set = set()
    result: List[str] = []
    for item in items:
        key = item.strip()
        if key and key not in seen:
            seen.add(key)
            result.append(key)
    return result


def bulleted_list(
    items: List[str],
    top_n: int = 10,
    header: Optional[str] = None,
) -> str:
    """Format a list of strings as a bullet-point block."""
    unique = deduplicated([x for x in items if str(x).strip()])
    if not unique:
        return "Unknown"

    shown, overflow = unique[:top_n], max(0, len(unique) - top_n)
    lines = ([header] if header else []) + [f"• {x}" for x in shown]
    if overflow:
        lines.append(f"... and {overflow} more.")
    return "\n".join(lines)


def best_name_from_path(path: str) -> str:
    """Extract a human-readable name from a file path."""
    try:
        parts = Path(path).parts
        return parts[-1] if parts else "Unknown"
    except Exception:
        return "Unknown"


def guess_system_application(file_path: str) -> str:
    """Infer a system/application name from a Semgrep result file path."""
    p = Path(file_path)
    parts = list(p.parts)

    for marker in ("testing", "GitHub"):
        if marker in parts:
            idx = parts.index(marker)
            if idx + 1 < len(parts):
                return parts[idx + 1]

    return p.parent.name or "Unknown"


def token_matches(value: str, token_set: frozenset) -> bool:
    """
    Return True if the normalised *value* exactly equals any token in *token_set*,
    or if any token in *token_set* is a whole-word substring of *value*.

    Uses exact-token matching rather than raw substring search to avoid false
    positives (e.g. "SECP" matching "ACCEPT", or "DH" matching "SHA").
    """
    normalised = value.upper().strip()
    if normalised in token_set:
        return True
    # Also allow tokens that appear as dash/space-delimited words within the value
    # (e.g. "RSA-OAEP" should match "RSA"; "ECDHE-RSA" should match "ECDH").
    parts = re.split(r"[-_/\s]+", normalised)
    return bool(token_set.intersection(parts))


def parse_key_length(key_length_str: str) -> Optional[int]:
    """
    Extract the first integer from a key-length string such as '128', '256 bits',
    'RSA-2048', etc.  Returns None if no number can be parsed.
    """
    if not key_length_str or str(key_length_str).strip().lower() in ("unknown", ""):
        return None
    match = re.search(r"\d+", str(key_length_str))
    return int(match.group()) if match else None


# ---------------------------------------------------------------------------
# CBOM sheet builders
# ---------------------------------------------------------------------------

def extract_line_number(result: Dict[str, Any]) -> Optional[int]:
    for location in (result.get("start"), (result.get("extra") or {}).get("line")):
        if isinstance(location, int):
            return location
        if isinstance(location, dict) and isinstance(location.get("line"), int):
            return location["line"]
    return None


def format_file_location(path: str, line: Optional[int]) -> str:
    if not path:
        return ""
    return f"{path} ({line})" if line is not None else path


def add_cbom_numbering(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    """Add a hierarchical '1.1', '1.2', … numbering column as the first column."""
    df = df.copy()
    if df.empty or group_col not in df.columns:
        df.insert(0, "# (CBOM)", pd.Series(dtype=str))
        return df
    group_ids = {
        name: idx + 1
        for idx, name in enumerate(df[group_col].dropna().unique())
    }
    group_num  = df[group_col].map(group_ids).fillna(0).astype(int)
    within_num = df.groupby(group_col, dropna=False).cumcount() + 1
    df.insert(0, "# (CBOM)", group_num.astype(str) + "." + within_num.astype(str))
    return df


def build_cbom_rows(semgrep: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for result in semgrep.get("results", []):
        path  = result.get("path", "") or ""
        line  = extract_line_number(result)
        extra = result.get("extra") or {}
        meta  = extra.get("metadata") or {}
        cbom  = meta.get("cbom") or {}

        rows.append({
            "System / Application":  coalesce(meta.get("system_application"), default=guess_system_application(path)),
            "Cryptographic Function": coalesce(cbom.get("cryptographic_function"), meta.get("cryptographic_function")),
            "Algorithm Used":         coalesce(cbom.get("algorithm_used"), meta.get("algorithm_name")),
            "Library / Module":       coalesce(cbom.get("library_module"), meta.get("library")),
            "File / Location":        format_file_location(path, line),
            "Key Length":             coalesce(cbom.get("key_length")),
            "Purpose / Usage":        coalesce(cbom.get("purpose_usage"), meta.get("purpose")),
            "Crypto-Agility Support": coalesce(meta.get("crypto_agility"), cbom.get("crypto_agility")),
        })

    return rows


def summarize_locations(locations: List[str], top_n: int = 3) -> str:
    unique = deduplicated([(loc or "").strip() for loc in locations])
    if not unique:
        return "Unknown"

    shown, overflow = unique[:top_n], max(0, len(unique) - top_n)
    lines = [f"• {x}" for x in shown]
    if overflow:
        lines.append(f"... and {overflow} more files.")
    return "\n".join(lines)


def build_cbom_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Collapse the full CBOM into one row per Cryptographic Function."""
    rows: List[Dict[str, Any]] = []

    for crypto_func, group in df.groupby("Cryptographic Function", dropna=False, sort=True):
        rows.append({
            "System / Application":   group["System / Application"].iloc[0],
            "Cryptographic Function":  str(crypto_func).strip() or "Unknown",
            "Algorithm Used":          group["Algorithm Used"].iloc[0],
            "Library / Module":        group["Library / Module"].iloc[0],
            "File / Location":         summarize_locations(group["File / Location"].tolist()),
            "Key Length":              group["Key Length"].iloc[0],
            "Purpose / Usage":         group["Purpose / Usage"].iloc[0],
            "Crypto-Agility Support":  group["Crypto-Agility Support"].iloc[0],
        })

    if not rows:
        return pd.DataFrame(columns=CBOM_COLS)

    df_sum = pd.DataFrame(rows)
    df_sum.sort_values(
        ["Cryptographic Function", "System / Application"],
        inplace=True,
        kind="stable",
    )
    df_sum = add_cbom_numbering(df_sum, group_col="Cryptographic Function")
    return df_sum[CBOM_COLS]


# ---------------------------------------------------------------------------
# SBOM / Grype sheet builder
# ---------------------------------------------------------------------------

def parse_syft_sbom(sbom: Dict[str, Any]) -> Dict[str, Any]:
    source      = sbom.get("source") or {}
    source_name = coalesce(source.get("name"), default="")
    system_name = (
        best_name_from_path(source_name) if source_name
        else coalesce(sbom.get("name"), default="Unknown")
    )

    modules: List[str] = []
    for artifact in (sbom.get("artifacts") or []):
        name, version, kind = artifact.get("name"), artifact.get("version"), artifact.get("type")
        if name and version:
            modules.append(f"{name} ({version})")
        elif name:
            modules.append(f"{name} ({kind})" if kind else name)

    developer = (
        "WordPress Foundation"
        if any("wordpress" in m.lower() for m in modules)
        else "Unknown Vendor"
    )

    return {
        "system_name":       system_name,
        "third_party_modules": modules,
        "developer_guess":   developer,
        "source_name":       source_name,
    }


def parse_grype(grype: Dict[str, Any]) -> Dict[str, Any]:
    vuln_lines: List[str] = []
    severities: List[str] = []

    for match in (grype.get("matches") or []):
        vuln     = match.get("vulnerability") or {}
        artifact = match.get("artifact") or {}

        cve      = coalesce(vuln.get("id"), default="UNKNOWN")
        severity = coalesce(vuln.get("severity"), default="UNKNOWN").upper()
        pkg      = coalesce(artifact.get("name"), default="package")
        version  = artifact.get("version")
        pkg_str  = f"{pkg} {version}".strip() if version else pkg

        vuln_lines.append(f"{pkg_str}: [{severity}] {cve}")
        severities.append(severity)

    return {
        "vuln_lines":    vuln_lines,
        "critical_level": max_severity(severities),
        "match_count":   len(vuln_lines),
    }


def build_sbom_sheet(
    df_cbom: pd.DataFrame,
    sbom: Dict[str, Any],
    grype: Dict[str, Any],
) -> pd.DataFrame:
    syft_data  = parse_syft_sbom(sbom)
    grype_data = parse_grype(grype)
    system     = syft_data["system_name"]

    cbom_systems = set(df_cbom["System / Application"].dropna().astype(str))
    link_to_cbom = (
        f"See CBOM entries for {system}" if system in cbom_systems
        else "See CBOM sheet"
    )

    software_component = "\n".join([
        "[CODE LEVEL]",
        f"App/System: {system}",
        f"SBOM Source: {syft_data.get('source_name') or 'Unknown'}",
    ])

    vuln_status = (
        "No known vulnerabilities found"
        if grype_data["match_count"] == 0
        else bulleted_list(grype_data["vuln_lines"], top_n=40, header="[ISSUES FOUND]")
    )

    row = {
        "#":                                        1,
        "System / Application":                     system,
        "Purpose / Usage":                          "",
        "URL":                                      "",
        "Services Mode":                            "",
        "Target Customer":                          "",
        "Software Component":                       software_component,
        "Vulnerability Status":                     vuln_status,
        "Third-party Modules":                      bulleted_list(syft_data["third_party_modules"], top_n=60, header="[DEPENDENCIES]"),
        "External APIs or Services":                "",
        "Critical Level":                           grype_data["critical_level"],
        "Data Category":                            "",
        "Is the application/system currently in use?": "",
        "Application/System Developer":             syft_data["developer_guess"],
        "Vendor's Name":                            "",
        "Does the agency have expertise?":          "",
        "Does the agency have a special budget allocation?": "",
        "Link to CBOM":                             link_to_cbom,
    }

    return pd.DataFrame([row], columns=SBOM_COLS)


# ---------------------------------------------------------------------------
# PQC Risk sheet
# ---------------------------------------------------------------------------

class PqcCategory:
    """Enumerated PQC risk categories for consistent handling."""
    SHOR          = "shor"           # Asymmetric — fully broken by quantum
    CLASSICALLY_BROKEN = "classical" # Already insecure without quantum
    GROVER_UPGRADE = "grover_upgrade" # Symmetric/hash — upgrade key size
    GROVER_ADEQUATE = "grover_ok"    # Already quantum-adequate key size
    UNKNOWN       = "unknown"        # Cannot classify


def categorise_algorithm(algo_used: str, crypto_func: str, key_length_str: str) -> PqcCategory:
    """
    Classify an algorithm into a PQC risk category.

    Priority order:
      1. Shor-vulnerable asymmetric algorithms (always Critical regardless of key length)
      2. Classically broken algorithms (flag as legacy, quantum is secondary)
      3. Grover-adequate (already safe at current size)
      4. Grover-needs-upgrade (safe with larger key/digest)
      5. Function-level fallback for asymmetric/symmetric functions
      6. Unknown
    """
    algo = algo_used.upper().strip()
    func = crypto_func.strip().lower()

    if token_matches(algo, SHOR_VULNERABLE_ALGORITHMS) or func in SHOR_VULNERABLE_FUNCTIONS:
        return PqcCategory.SHOR

    if token_matches(algo, CLASSICALLY_BROKEN_ALGORITHMS):
        return PqcCategory.CLASSICALLY_BROKEN

    if token_matches(algo, GROVER_ADEQUATE_ALGORITHMS):
        return PqcCategory.GROVER_ADEQUATE

    if token_matches(algo, GROVER_NEEDS_UPGRADE_ALGORITHMS):
        # Also check key length: if AES with key_length >= 256, it's adequate
        key_bits = parse_key_length(key_length_str)
        if key_bits is not None and key_bits >= 256:
            return PqcCategory.GROVER_ADEQUATE
        return PqcCategory.GROVER_UPGRADE

    # Function-level fallback
    if func in SYMMETRIC_FUNCTIONS or func in HASH_FUNCTIONS:
        return PqcCategory.GROVER_UPGRADE

    return PqcCategory.UNKNOWN


def classify_pqc(
    algo_used: str,
    crypto_func: str,
    key_length_str: str = "Unknown",
) -> Tuple[str, str]:
    """
    Return (Severity, Risk description) for use in the PQC_Risk sheet.

    Severity scale:
      Critical  — Fully broken by a Cryptographically Relevant Quantum Computer (CRQC)
                  via Shor's algorithm. Migrate immediately.
      High      — Classically broken (no quantum computer needed). Migrate immediately.
      Medium    — Quantum-weakened; current key size should be increased.
      Low       — Already meets post-quantum security at current key/digest size.
      Info      — Cannot classify; manual review required.
    """
    category = categorise_algorithm(algo_used, crypto_func, key_length_str)

    if category == PqcCategory.SHOR:
        return (
            "Critical",
            f"Fully broken by Shor's algorithm on a CRQC. {algo_used} provides no "
            "long-term security. Migrate to NIST PQC standards (ML-KEM / ML-DSA / SLH-DSA).",
        )

    if category == PqcCategory.CLASSICALLY_BROKEN:
        return (
            "High",
            f"{algo_used} is already classically broken and must be replaced regardless "
            "of quantum threats. Quantum risk is a secondary concern here.",
        )

    if category == PqcCategory.GROVER_UPGRADE:
        return (
            "Medium",
            f"Grover's algorithm halves the effective bit-security of {algo_used}. "
            "Increase key/digest size (e.g. AES-128 → AES-256, SHA-256 → SHA-384+). "
            "NIST does not require full migration away from symmetric primitives.",
        )

    if category == PqcCategory.GROVER_ADEQUATE:
        return (
            "Low",
            f"{algo_used} already provides ≥128-bit post-quantum security at its current "
            "key/digest size. No immediate PQC migration action required by NIST guidance.",
        )

    # UNKNOWN
    return (
        "Info",
        f"Unable to confidently classify {algo_used} for PQC risk. "
        "Manual review recommended to confirm algorithm usage and select a migration path.",
    )


def infer_asset_type(system_app: str) -> str:
    return "Operating System" if is_ipv4(system_app.strip()) else "Application Code"


def build_pqc_risk_sheet(df_cbom_summary: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for _, row in df_cbom_summary.iterrows():
        system      = coalesce(row.get("System / Application"), default="Unknown")
        algo        = coalesce(row.get("Algorithm Used"), default="Unknown")
        crypto_func = coalesce(row.get("Cryptographic Function"), default="Unknown")
        key_length  = coalesce(row.get("Key Length"), default="Unknown")

        severity, risk = classify_pqc(algo, crypto_func, key_length)

        rows.append({
            "System / Application": system,
            "Asset Type":           infer_asset_type(system),
            "Algorithm Used":       algo,
            "Purpose / Usage":      crypto_func,
            "Severity":             severity,
            "Risk":                 risk,
            "Risk Owner":           "IT Security Team",
        })

    severity_order = {"CRITICAL": 1, "HIGH": 2, "MEDIUM": 3, "LOW": 4, "INFO": 5}
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=PQC_RISK_COLS)
    df["__rank"] = df["Severity"].str.upper().map(severity_order).fillna(99).astype(int)
    df.sort_values(
        ["System / Application", "__rank", "Purpose / Usage", "Algorithm Used"],
        inplace=True,
        kind="stable",
    )
    df.drop(columns=["__rank"], inplace=True)
    df.insert(0, "#", range(1, len(df) + 1))
    return df[PQC_RISK_COLS]


# ---------------------------------------------------------------------------
# PQC 5×5 Risk Matrix sheet
# ---------------------------------------------------------------------------

def risk_level_from_score(score: int) -> str:
    """Map a numeric risk score to a qualitative level using a standard 5×5 matrix."""
    if score >= 15:
        return "Very High"
    if score >= 10:
        return "High"
    if score >= 5:
        return "Medium"
    return "Low"


def evaluate_pqc_matrix(
    algo_used: str,
    crypto_func: str,
    key_length_str: str = "Unknown",
) -> Tuple[str, str, int, int, str, str]:
    """
    Return (Risk, Root Cause, Impact, Likelihood, Existing Controls, Mitigation Plan)
    for the 5×5 PQC Risk Matrix sheet.

    Scoring rationale (aligned with NIST PQC guidance):
    ┌─────────────────────────────┬────────┬────────────┬──────────────────────────────┐
    │ Category                    │ Impact │ Likelihood │ Notes                        │
    ├─────────────────────────────┼────────┼────────────┼──────────────────────────────┤
    │ Shor (asymmetric)           │   5    │     5      │ Fully broken; migrate now    │
    │ Classically broken          │   5    │     5      │ Already exploitable today    │
    │ Grover – needs upgrade      │   3    │     3      │ Weakened; manageable         │
    │ Grover – already adequate   │   1    │     2      │ Low residual risk            │
    │ Unknown                     │   3    │     2      │ Conservative estimate        │
    └─────────────────────────────┴────────┴────────────┴──────────────────────────────┘
    """
    controls = "Standard IT Security Controls"
    category = categorise_algorithm(algo_used, crypto_func, key_length_str)

    if category == PqcCategory.SHOR:
        return (
            "Exposure to Shor's Algorithm (Quantum Integer Factorization / Discrete Logarithm)",
            f"Usage of asymmetric algorithm ({algo_used}) which will be fully broken "
            "by a Cryptographically Relevant Quantum Computer (CRQC)",
            5, 5,
            controls,
            "Migrate to NIST-approved PQC algorithms: ML-KEM (FIPS 203) for key "
            "establishment, ML-DSA (FIPS 204) or SLH-DSA (FIPS 205) for signatures. "
            "Prioritise long-lived data encrypted today (harvest-now-decrypt-later threat).",
        )

    if category == PqcCategory.CLASSICALLY_BROKEN:
        return (
            "Classically Broken Algorithm (no quantum computer required)",
            f"{algo_used} has known classical attacks (collision/preimage/key-recovery) "
            "that make it insecure today; quantum risk is a secondary concern",
            5, 5,
            controls,
            f"Replace {algo_used} immediately with a modern algorithm (AES-256, "
            "SHA-256+, or a NIST PQC primitive). Do not wait for a CRQC.",
        )

    if category == PqcCategory.GROVER_UPGRADE:
        key_bits = parse_key_length(key_length_str)
        if key_bits:
            post_quantum_bits = key_bits // 2
            key_note = (
                f" Current key/digest is {key_bits}-bit "
                f"(~{post_quantum_bits}-bit post-quantum security)."
            )
        else:
            key_note = " Key/digest size unspecified — verify before assessing residual risk."

        if crypto_func.strip().lower() in HASH_FUNCTIONS:
            algo_kind, upgrade_advice = "hash", "SHA-256 → SHA-384 or SHA-512"
        else:
            algo_kind, upgrade_advice = "symmetric", "AES-128 → AES-256"

        return (
            "Exposure to Grover's Algorithm (Quantum Search – halves effective bit-security)",
            f"Usage of {algo_kind} algorithm ({algo_used}) whose effective security is "
            f"halved by Grover's algorithm.{key_note}",
            3, 3,
            controls,
            f"Increase key/digest size: {upgrade_advice}. "
            "NIST does not require migration away from symmetric primitives — "
            "doubling key length restores post-quantum security.",
        )

    if category == PqcCategory.GROVER_ADEQUATE:
        return (
            "Residual Quantum Risk – Low (adequate key size for post-quantum security)",
            f"{algo_used} already provides ≥128-bit post-quantum security at its current "
            "key/digest size; Grover's algorithm does not reduce it below acceptable thresholds",
            1, 2,
            controls,
            "No immediate migration action required per NIST guidance. "
            "Continue monitoring NIST PQC standards for any future algorithm deprecations.",
        )

    # UNKNOWN
    return (
        "PQC Exposure Requires Manual Review",
        f"Algorithm ({algo_used}) could not be confidently mapped to a PQC risk category; "
        "function context or key length may be missing",
        3, 2,
        controls,
        "Perform crypto discovery validation, confirm algorithm and key length, "
        "then select the appropriate NIST PQC migration path.",
    )


def build_pqc_risk_matrix_sheet(df_cbom_summary: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for _, row in df_cbom_summary.iterrows():
        system      = coalesce(row.get("System / Application"), default="Unknown")
        algo        = coalesce(row.get("Algorithm Used"), default="Unknown")
        crypto_func = coalesce(row.get("Cryptographic Function"), default="Unknown")
        key_length  = coalesce(row.get("Key Length"), default="Unknown")

        risk, root, impact, likelihood, controls, mitigation = evaluate_pqc_matrix(
            algo, crypto_func, key_length
        )
        score = impact * likelihood

        rows.append({
            "System/Hardware/Software Name": system,
            "Cryptographic Algorithm":       algo,
            "Risk":                          risk,
            "Root Cause":                    root,
            "Impact":                        impact,
            "Likelihood":                    likelihood,
            "Risk Score":                    score,
            "Risk Level":                    risk_level_from_score(score),
            "Existing Controls":             controls,
            "Mitigation Plan":               mitigation,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=PQC_MATRIX_COLS)
    if not df.empty:
        df.sort_values(
            ["Risk Score", "System/Hardware/Software Name", "Cryptographic Algorithm"],
            ascending=[False, True, True],
            inplace=True,
            kind="stable",
        )
    df.insert(0, "#", range(1, len(df) + 1))
    return df[PQC_MATRIX_COLS]



# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# MiniPQC sheet helpers  (4 sheets mirroring SBOM/CBOM/PQC schemas)
# ---------------------------------------------------------------------------

# Severity integer → label mapping (matches scanner output)
_SEV_LABEL: Dict[int, str] = {
    0: "Info",
    1: "Low",
    2: "Medium",
    3: "High",
    4: "Very High",
    5: "Critical",
}

# Module ID → human name
_MODULE_NAMES: Dict[int, str] = {
    1:  "Environment",
    2:  "Firmware / Boot",
    3:  "Kernel",
    4:  "IPsec",
    5:  "Nginx",
    6:  "OpenSSH",
    7:  "OpenVPN",
    8:  "Email (Postfix)",
    9:  "GnuPG",
    10: "Language Runtimes",
    11: "PKI / CA Tools",
    14: "OpenSSL / GnuTLS",
    15: "Apache",
    16: "WireGuard",
}

# Findings with Kind==1 are "component findings"; Kind==0 are recommendations.
# We use Kind==1 items to infer installed components and Kind==0 for advice.

def _module_name(module_id: int) -> str:
    return _MODULE_NAMES.get(module_id, f"Module {module_id}")


def _sev_label(sev_int: int) -> str:
    return _SEV_LABEL.get(sev_int, "Info")


def _extract_system_ip(data: Dict[str, Any]) -> str:
    """Best-effort extraction of the scanned host identifier."""
    return data.get("server_ip") or data.get("hostname") or "Unknown"


def _collect_components(recs: List[Dict[str, Any]]) -> List[str]:
    """Return bullet lines for the Software Component cell (Kind==1 findings)."""
    lines: List[str] = []
    seen: set = set()
    for item in recs:
        if item.get("Kind") == 1:
            text = (item.get("Text") or "").strip()
            if text and text not in seen:
                seen.add(text)
                lines.append(f"• {text}")
    return lines


def _collect_third_party(recs: List[Dict[str, Any]]) -> List[str]:
    """Return third-party module / library bullet lines."""
    lines: List[str] = []
    seen: set = set()
    keywords = ("openssl", "gnutls", "oqs", "bouncy", "libssl", "libcrypto",
                 "python", "java", "node", "npm", "pip", "pyca", "pycrypto",
                 "paramiko", "pyopenssl", "cryptography", "nacl", "gnupg",
                 "openvpn", "openssh", "nginx", "apache", "wireshark", "tshark")
    for item in recs:
        text = (item.get("Text") or "").strip().lower()
        for kw in keywords:
            if kw in text and text not in seen:
                seen.add(text)
                lines.append(f"• {item['Text'].strip()}")
                break
    return lines


def _vuln_status(recs: List[Dict[str, Any]]) -> str:
    """Summarise vulnerability / issue status from high-severity recommendations."""
    issues = [
        f"• {item.get('Text','').strip()}"
        for item in recs
        if item.get("Kind") == 0 and item.get("Severity", 0) >= 4
    ]
    if not issues:
        return "No known vulnerabilities found (OS Patching Managed Separately)"
    return "[ISSUES FOUND]\n" + "\n".join(issues[:20])


# ── Sheet 1 : MiniPQC_SBOM  (mirrors SBOM_Grype layout) ────────────────────

MINIPQC1_COLS = [
    "#",
    "System / Application",
    "Purpose / Usage",
    "URL",
    "Services Mode",
    "Target Customer",
    "Software Component",
    "Vulnerability Status",
    "Third-party Modules",
    "External APIs or Services",
    "Critical Level",
    "Data Category",
    "Is the application/system currently in use?",
    "Application/System Developer",
    "Vendor's Name",
    "Does the agency have expertise?",
    "Does the agency have a special budget allocation?",
    "Link to CBOM",
]

def build_minipqc1(data: Dict[str, Any]) -> pd.DataFrame:
    """MiniPQC_SBOM — one row per scanned host, mirroring SBOM_Grype."""
    recs   = data.get("recommendations", [])
    system = _extract_system_ip(data)

    component_lines: List[str] = ["[OS/SERVICE LEVEL]"]
    component_lines += _collect_components(recs)
    software_component = "\n".join(component_lines)

    # Determine highest severity seen
    max_sev = max((item.get("Severity", 0) for item in recs), default=0)
    critical_level = _sev_label(max_sev)

    row = {
        "#":                                          1,
        "System / Application":                       system,
        "Purpose / Usage":                            "",
        "URL":                                        "",
        "Services Mode":                              "",
        "Target Customer":                            "",
        "Software Component":                         software_component,
        "Vulnerability Status":                       _vuln_status(recs),
        "Third-party Modules":                        "\n".join(_collect_third_party(recs)),
        "External APIs or Services":                  "",
        "Critical Level":                             critical_level,
        "Data Category":                              "",
        "Is the application/system currently in use?": "",
        "Application/System Developer":               "",
        "Vendor's Name":                              "",
        "Does the agency have expertise?":            "",
        "Does the agency have a special budget allocation?": "",
        "Link to CBOM":                               "See MiniPQC_CBOM sheet",
    }
    return pd.DataFrame([row], columns=MINIPQC1_COLS)


# ── Sheet 2 : MiniPQC_CBOM  (mirrors CBOM layout with Algorithm Category) ──

MINIPQC2_COLS = [
    "# (CBOM)",
    "System / Application",
    "Cryptographic Function",
    "Algorithm Used",
    "Algorithm Category",
    "Library / Module",
    "File / Location",
    "Key Length",
    "Purpose / Usage",
    "Crypto-Agility Support",
]

# Map module → default crypto function label for algorithm items
_FUNC_HINT: Dict[str, str] = {
    "Kernel":           "Kernel Primitive",
    "OpenSSH":          "Key Exchange / Authentication",
    "OpenSSL / GnuTLS": "TLS / PKI",
    "Nginx":            "TLS",
    "Apache":           "TLS",
    "OpenVPN":          "VPN Tunnel",
    "Language Runtimes":"Library Call",
    "PKI / CA Tools":   "Certificate",
}

def _algo_category(algo: str, func: str) -> str:
    """Quick label: Asymmetric / Symmetric / Hash / PQC / Unknown."""
    a = algo.upper()
    f = func.lower()
    if any(x in a for x in ("RSA", "ECDH", "ECDSA", "DH", "ED25519", "ED448", "DSA")):
        return "Asymmetric"
    if any(x in a for x in ("ML-KEM", "ML-DSA", "SLH-DSA", "KYBER", "DILITHIUM", "SPHINCS")):
        return "PQC"
    if any(x in a for x in ("AES", "CHACHA", "3DES", "DES", "RC4", "BLOWFISH")):
        return "Symmetric"
    if any(x in a for x in ("SHA", "MD5", "BLAKE", "HMAC", "RIPEMD", "WHIRLPOOL")):
        return "Hash"
    if "asymmetric" in f or "signature" in f or "key exchange" in f:
        return "Asymmetric"
    if "hash" in f:
        return "Hash"
    if "symmetric" in f or "encrypt" in f:
        return "Symmetric"
    return "Unknown"


def _agility(algo: str) -> str:
    a = algo.upper()
    if any(x in a for x in ("ML-KEM", "ML-DSA", "SLH-DSA")):
        return "High (PQC Native)"
    if any(x in a for x in ("AES-256", "SHA-384", "SHA-512", "SHA3-384", "SHA3-512")):
        return "High (Quantum-Adequate)"
    return "Low (Classical)"


def build_minipqc2(data: Dict[str, Any]) -> pd.DataFrame:
    """MiniPQC_CBOM — one row per crypto algorithm detected, mirroring CBOM."""
    recs   = data.get("recommendations", [])
    system = _extract_system_ip(data)
    rows: List[Dict[str, Any]] = []

    # Collect algorithm mentions from Kind==1 findings
    algo_keywords = re.compile(
        r"\b(RSA[\w\s(#+\-]*?(?=\s*[,;|\n]|$)|ECDH[\w\-]+|ECDSA[\w\-]*|"
        r"ED25519|ED448|AES-\d+|AES\b|BLAKE2[A-Z]-\d+|BLAKE3\b|"
        r"SHA-\d+|SHA\d+|SHA3-\d+|MD5\b|HMAC-SHA\d*|"
        r"ML-KEM-\d+|ML-DSA-\d+|SLH-DSA[\w\-]*|"
        r"CHACHA20[\w\-]*|POLY1305|DH\b|DSA\b)",
        re.IGNORECASE,
    )

    seen_algos: set = set()
    group_counter: Dict[str, int] = {}
    item_counter: Dict[str, int] = {}

    for item in recs:
        if item.get("Kind") != 1:
            continue
        text    = item.get("Text", "")
        details = item.get("Details", "")
        module  = _module_name(item.get("ModuleID", 0))
        lib_loc = text.strip()
        func    = _FUNC_HINT.get(module, "Cryptographic Function")

        # Try to find explicit algorithm tokens in the text
        matches = algo_keywords.findall(text + " " + details)
        algos   = [m.strip() for m in matches if m.strip()] if matches else []

        if not algos:
            continue

        for algo in algos:
            key = (module, algo.upper())
            if key in seen_algos:
                continue
            seen_algos.add(key)

            cat = _algo_category(algo, func)
            if module not in group_counter:
                group_counter[module] = len(group_counter) + 1
                item_counter[module]  = 0
            item_counter[module] += 1

            rows.append({
                "_grp":   group_counter[module],
                "_item":  item_counter[module],
                "System / Application":  system,
                "Cryptographic Function": func,
                "Algorithm Used":         algo,
                "Algorithm Category":     cat,
                "Library / Module":       f"[{module}] {lib_loc}"[:120],
                "File / Location":        "System Kernel / Memory" if module == "Kernel" else module,
                "Key Length":             "Variable",
                "Purpose / Usage":        "System Security",
                "Crypto-Agility Support": _agility(algo),
            })

    if not rows:
        return pd.DataFrame(columns=MINIPQC2_COLS)

    df = pd.DataFrame(rows)
    df.sort_values(["_grp", "_item"], inplace=True, kind="stable")
    df["# (CBOM)"] = df["_grp"].astype(str) + "." + df["_item"].astype(str)
    return df[MINIPQC2_COLS]


# ── Sheet 3 : MiniPQC_Risk  (mirrors PQC_Risk, column headers in Malay) ────

MINIPQC3_COLS = [
    "#",
    "Nama Sistem/ Perkakasan/Perisian",
    "Jenis Aset",
    "Algoritma Kriptografi",
    "Kegunaan Algoritma Kriptografi",
    "Tahap Kritikal",
    "Risiko",
    "Pemilik Risiko",
]

_SEVERITY_MALAY: Dict[str, str] = {
    "Critical": "Kritikal",
    "High":     "Tinggi",
    "Medium":   "Sederhana",
    "Low":      "Rendah",
    "Info":     "Maklumat",
}

_ASSET_TYPE_MALAY: Dict[str, str] = {
    "Operating System": "Operating System",
    "Application Code": "Application Code",
}


def build_minipqc3(data: Dict[str, Any]) -> pd.DataFrame:
    """MiniPQC_Risk — PQC risk per algorithm, Malay column headers."""
    df2  = build_minipqc2(data)
    system = _extract_system_ip(data)
    rows: List[Dict[str, Any]] = []

    for _, row in df2.iterrows():
        algo       = row["Algorithm Used"]
        func       = row["Cryptographic Function"]
        key_len    = row.get("Key Length", "Unknown")
        sev_en, _  = classify_pqc(algo, func, key_len)
        sev_my     = _SEVERITY_MALAY.get(sev_en, sev_en)
        asset_type = "Operating System" if is_ipv4(system) else "Application Code"
        _, risk_en = classify_pqc(algo, func, key_len)

        rows.append({
            "Nama Sistem/ Perkakasan/Perisian": system,
            "Jenis Aset":                       asset_type,
            "Algoritma Kriptografi":            algo,
            "Kegunaan Algoritma Kriptografi":   func,
            "Tahap Kritikal":                   sev_my,
            "Risiko":                           risk_en,
            "Pemilik Risiko":                   "IT Security Team",
        })

    if not rows:
        return pd.DataFrame(columns=MINIPQC3_COLS)

    df = pd.DataFrame(rows)
    # Sort by severity (Critical first)
    sev_order_my = {"Kritikal": 1, "Tinggi": 2, "Sederhana": 3, "Rendah": 4, "Maklumat": 5}
    df["__rank"] = df["Tahap Kritikal"].map(sev_order_my).fillna(9)
    df.sort_values(["__rank", "Nama Sistem/ Perkakasan/Perisian", "Algoritma Kriptografi"],
                   inplace=True, kind="stable")
    df.drop(columns=["__rank"], inplace=True)
    df.insert(0, "#", range(1, len(df) + 1))
    return df[MINIPQC3_COLS]


# ── Sheet 4 : MiniPQC_RiskMatrix  (mirrors PQC_RiskMatrix, Malay headers) ──

MINIPQC4_COLS = [
    "#",
    "Nama Sistem/ Perkakasan/Perisian",
    "Algoritma Kriptografi",
    "Risiko",
    "Punca Risiko",
    "Impak",
    "Kemungkinan (Likelihood)",
    "Skor Risiko",
    "Risk Level",
    "Kawalan Sedia Ada",
    "Pelan Mitigasi",
]

_RISK_LEVEL_MALAY: Dict[str, str] = {
    "Very High": "Sangat Tinggi",
    "High":      "Tinggi",
    "Medium":    "Sederhana",
    "Low":       "Rendah",
}


def build_minipqc4(data: Dict[str, Any]) -> pd.DataFrame:
    """MiniPQC_RiskMatrix — 5×5 matrix per algorithm, Malay column headers."""
    df2    = build_minipqc2(data)
    system = _extract_system_ip(data)
    rows: List[Dict[str, Any]] = []

    for _, row in df2.iterrows():
        algo    = row["Algorithm Used"]
        func    = row["Cryptographic Function"]
        key_len = row.get("Key Length", "Unknown")

        risk, root, impact, likelihood, controls, mitigation = evaluate_pqc_matrix(
            algo, func, key_len
        )
        score      = impact * likelihood
        level_en   = risk_level_from_score(score)
        level_my   = _RISK_LEVEL_MALAY.get(level_en, level_en)

        rows.append({
            "Nama Sistem/ Perkakasan/Perisian": system,
            "Algoritma Kriptografi":            algo,
            "Risiko":                           risk,
            "Punca Risiko":                     root,
            "Impak":                            impact,
            "Kemungkinan (Likelihood)":         likelihood,
            "Skor Risiko":                      score,
            "Risk Level":                       level_my,
            "Kawalan Sedia Ada":                controls,
            "Pelan Mitigasi":                   mitigation,
        })

    if not rows:
        return pd.DataFrame(columns=MINIPQC4_COLS)

    df = pd.DataFrame(rows)
    df.sort_values(["Skor Risiko", "Nama Sistem/ Perkakasan/Perisian", "Algoritma Kriptografi"],
                   ascending=[False, True, True], inplace=True, kind="stable")
    df.insert(0, "#", range(1, len(df) + 1))
    return df[MINIPQC4_COLS]


# Keep legacy aliases so references in main() still resolve
MINIPQC_COLS         = MINIPQC1_COLS
MINIPQC_SUMMARY_COLS = MINIPQC2_COLS
build_minipqc_sheet    = build_minipqc1
build_minipqc_summary  = build_minipqc2


# Excel styling helpers
# ---------------------------------------------------------------------------

def apply_header_style(worksheet) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font   = Font(bold=True)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(**{side: Side(style="thin") for side in ("left", "right", "top", "bottom")})

    for cell in worksheet[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
        cell.border    = border


def autofit_columns(worksheet, max_width: int = 90) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx, col_cells in enumerate(worksheet.columns, start=1):
        max_len = max(
            (
                max((len(line) for line in str(cell.value).splitlines()), default=0)
                if cell.value is not None else 0
                for cell in col_cells
            ),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max_width, max(10, max_len + 2)
        )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _apply_severity_colours(ws, severity_col_idx: int) -> None:
    """Colour-code severity cells by value (1-indexed column position)."""
    from openpyxl.styles import Font, PatternFill

    palette = {
        "Critical":  ("FFFFFF", "C00000"),
        "Very High": ("FFFFFF", "E04000"),
        "High":      ("FFFFFF", "FF0000"),
        "Medium":    ("000000", "FFBF00"),
        "Low":       ("000000", "92D050"),
        "Info":      ("000000", "D9D9D9"),
    }
    for row in ws.iter_rows(min_row=2):
        cell = row[severity_col_idx - 1]
        colours = palette.get(str(cell.value))
        if colours:
            fg, bg = colours
            cell.font = Font(bold=True, color=fg)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def main() -> None:
    # -- Load inputs ----------------------------------------------------------
    semgrep  = json.loads(INPUT_CBOM.read_text(encoding="utf-8"))
    sbom     = json.loads(INPUT_SBOM.read_text(encoding="utf-8"))
    grype    = json.loads(INPUT_GRYPE.read_text(encoding="utf-8"))
    minipqc  = json.loads(INPUT_MINIPQC.read_text(encoding="utf-8")) \
               if INPUT_MINIPQC.exists() else {}

    # -- Build CBOM sheets ----------------------------------------------------
    df_cbom_rows = pd.DataFrame(build_cbom_rows(semgrep))

    if df_cbom_rows.empty:
        df_cbom_rows = pd.DataFrame(columns=[c for c in CBOM_COLS if c != "# (CBOM)"])
    else:
        df_cbom_rows.sort_values(
            ["Cryptographic Function", "System / Application", "File / Location", "Algorithm Used"],
            inplace=True, kind="stable", na_position="last",
        )

    df_cbom_rows = add_cbom_numbering(df_cbom_rows, group_col="System / Application")
    for col in CBOM_COLS:
        if col not in df_cbom_rows.columns:
            df_cbom_rows[col] = None
    df_cbom_rows = df_cbom_rows[CBOM_COLS]
    df_cbom_sum  = build_cbom_summary(df_cbom_rows.drop(columns=["# (CBOM)"], errors="ignore"))

    # -- Build SBOM / Grype sheet ---------------------------------------------
    df_sbom = build_sbom_sheet(df_cbom_rows, sbom, grype)

    # -- Build PQC sheets (derived from CBOM summary) -------------------------
    df_pqc_risk   = build_pqc_risk_sheet(df_cbom_sum)
    df_pqc_matrix = build_pqc_risk_matrix_sheet(df_cbom_sum)

    # -- Build MiniPQC sheets (derived from minipqc.json) ---------------------
    df_mpqc1 = build_minipqc1(minipqc)   # SBOM-style inventory
    df_mpqc2 = build_minipqc2(minipqc)   # CBOM-style algorithm list
    df_mpqc3 = build_minipqc3(minipqc)   # PQC risk (Malay headers)
    df_mpqc4 = build_minipqc4(minipqc)   # Risk matrix (Malay headers)

    # -- Write workbook -------------------------------------------------------
    sheets = {
        "CBOM":              df_cbom_rows,
        "CBOM_Summary":      df_cbom_sum,
        "SBOM_Grype":        df_sbom,
        "PQC_Risk":          df_pqc_risk,
        "PQC_RiskMatrix":    df_pqc_matrix,
        "MiniPQC_SBOM":      df_mpqc1,
        "MiniPQC_CBOM":      df_mpqc2,
        "MiniPQC_Risk":      df_mpqc3,
        "MiniPQC_RiskMatrix": df_mpqc4,
    }

    # Severity column indices (1-based) for colour-coding
    severity_col_map: Dict[str, int] = {
        "PQC_Risk":           PQC_RISK_COLS.index("Severity") + 1,
        "MiniPQC_Risk":       MINIPQC3_COLS.index("Tahap Kritikal") + 1,
        "MiniPQC_RiskMatrix": MINIPQC4_COLS.index("Risk Level") + 1,
    }

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        for sheet_name, ws in writer.sheets.items():
            ws.freeze_panes = "A2"
            apply_header_style(ws)
            autofit_columns(ws)
            if sheet_name in severity_col_map:
                _apply_severity_colours(ws, severity_col_map[sheet_name])

    # -- Summary --------------------------------------------------------------
    print(f"[OK] Output: {OUTPUT_XLSX.resolve()}")
    print(
        "[OK] Rows — "
        + " | ".join(f"{name}={len(df)}" for name, df in sheets.items())
    )


if __name__ == "__main__":
    main()